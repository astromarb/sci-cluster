from __future__ import annotations

import json
import os
from dataclasses import replace
from pathlib import Path

import pytest

pd = pytest.importorskip("pandas")
openpyxl = pytest.importorskip("openpyxl")

import rmelts_mc_pipeline as rmp


def _minimal_prepared_df():
    row = {c: 0.0 for c in rmp.MELTS_OXIDE_ROWS}
    row.update(
        {
            "sample_label": "S1",
            "SiO2": 75.0,
            "TiO2": 0.05,
            "Al2O3": 13.0,
            "Fe2O3": 0.2,
            "FeO": 0.5,
            "MnO": 0.03,
            "MgO": 0.08,
            "CaO": 0.6,
            "Na2O": 4.2,
            "K2O": 4.0,
            "P2O5": 0.01,
            "H2O": 2.0,
        }
    )
    return pd.DataFrame([row])[rmp.PREPARED_MC_COLUMNS]


def test_mc_to_csv_rmelts_converts_and_splits_iron(tmp_path):
    mc_df = pd.DataFrame(
        [
            {
                "SampleID": "dup",
                "SiO2": 75.0,
                "TiO2": 0.05,
                "Al2O3": 13.0,
                "FeOt": 1.0,
                "MnO": 0.03,
                "MgO": 0.08,
                "CaO": 0.6,
                "Na2O": 4.2,
                "K2O": 4.0,
                "P2O5": 0.01,
                "H2O": 2.0,
            },
            {
                "SampleID": "dup",
                "SiO2": 74.8,
                "TiO2": 0.04,
                "Al2O3": 13.1,
                "FeOt": 0.9,
                "MnO": 0.03,
                "MgO": 0.07,
                "CaO": 0.55,
                "Na2O": 4.1,
                "K2O": 4.2,
                "P2O5": 0.01,
                "H2O": 2.0,
            },
        ]
    )
    mc_path = tmp_path / "mc.csv"
    mc_df.to_csv(mc_path, index=False)

    result = rmp.MC_to_csv_rMELTS(
        mc_path,
        output_dir=tmp_path,
        dataset_name="demo dataset",
        sample_id_col="SampleID",
        fe_total_col="FeOt",
        fe3fet=0.2,
        validate_only=True,
    )

    prepared = pd.read_csv(result.prepared_mc_csv_path)
    qc = pd.read_csv(result.qc_report_csv_path)

    assert result.num_samples == 2
    assert prepared["sample_label"].tolist() == ["dup", "dup__2"]
    feo_expected = 1.0 * (1.0 - 0.2)
    fe2o3_expected = 1.0 * 0.2 * (rmp.M_FE2O3 / (2.0 * rmp.M_FEO))
    row0 = prepared.iloc[0]
    assert row0["FeO"] == pytest.approx(feo_expected, rel=1e-12)
    assert row0["Fe2O3"] == pytest.approx(fe2o3_expected, rel=1e-12)
    assert "FeOt_original" in qc.columns
    assert set(prepared.columns) == set(rmp.PREPARED_MC_COLUMNS)


def test_mc_to_csv_rmelts_refuses_implicit_feot_repartition(tmp_path):
    mc_df = pd.DataFrame(
        [
            {
                "SampleID": "s1",
                "SiO2": 75.0,
                "TiO2": 0.05,
                "Al2O3": 13.0,
                "FeOt": 1.0,
                "MnO": 0.03,
                "MgO": 0.08,
                "CaO": 0.6,
                "Na2O": 4.2,
                "K2O": 4.0,
                "P2O5": 0.01,
                "H2O": 2.0,
            }
        ]
    )
    mc_path = tmp_path / "mc.csv"
    mc_df.to_csv(mc_path, index=False)

    with pytest.raises(rmp.RMeltsPipelineError):
        rmp.MC_to_csv_rMELTS(
            mc_path,
            output_dir=tmp_path,
            dataset_name="preserve_fe",
            sample_id_col="SampleID",
        )


def test_mc_to_csv_rmelts_preserves_direct_fe_speciation_without_fe_split(tmp_path):
    mc_df = pd.DataFrame(
        [
            {
                "SampleID": "s1",
                "SiO2": 75.0,
                "TiO2": 0.05,
                "Al2O3": 13.0,
                "FeOt": 0.66,  # traceability only; should not override direct Fe species
                "FeO": 0.49,
                "Fe2O3": 0.18,
                "MnO": 0.03,
                "MgO": 0.08,
                "CaO": 0.6,
                "Na2O": 4.2,
                "K2O": 4.0,
                "P2O5": 0.01,
                "H2O": 2.0,
            }
        ]
    )
    mc_path = tmp_path / "mc.csv"
    mc_df.to_csv(mc_path, index=False)

    result = rmp.MC_to_csv_rMELTS(
        mc_path,
        output_dir=tmp_path,
        dataset_name="preserve_direct_fe",
        sample_id_col="SampleID",
    )
    prepared = pd.read_csv(result.prepared_mc_csv_path)
    qc = pd.read_csv(result.qc_report_csv_path)
    assert prepared.loc[0, "FeO"] == pytest.approx(0.49)
    assert prepared.loc[0, "Fe2O3"] == pytest.approx(0.18)
    assert bool(qc.loc[0, "fe_split_applied"]) is False
    assert qc.loc[0, "FeOt_original"] == pytest.approx(0.66)


def test_mc_to_csv_rmelts_leaves_missing_optional_oxides_as_nan(tmp_path):
    mc_df = pd.DataFrame(
        [
            {
                "SampleID": "s1",
                "SiO2": 75.0,
                "TiO2": 0.05,
                "Al2O3": 13.0,
                "FeO": 0.49,
                "Fe2O3": 0.18,
                "MnO": 0.03,
                "MgO": 0.08,
                "CaO": 0.6,
                "Na2O": 4.2,
                "K2O": 4.0,
                "P2O5": 0.01,
                "H2O": 2.0,
                # intentionally omit Cr2O3 / NiO / CoO / CO2 / SO3 / halogens
            }
        ]
    )
    mc_path = tmp_path / "mc.csv"
    mc_df.to_csv(mc_path, index=False)

    result = rmp.MC_to_csv_rMELTS(
        mc_path,
        output_dir=tmp_path,
        dataset_name="missing_optional_nan",
        sample_id_col="SampleID",
    )
    prepared = pd.read_csv(result.prepared_mc_csv_path)
    assert pd.isna(prepared.loc[0, "Cr2O3"])
    assert pd.isna(prepared.loc[0, "NiO"])
    assert pd.isna(prepared.loc[0, "CoO"])

    params = rmp.MELTSRunParams(
        T1=1100, T2=700, dT=1, P1=400, P2=200, dP=25,
        fO2_constraint="TRUE", fO2_buffer="NNO", fO2_offset=0.0
    )
    wide = rmp._build_melts_input_wide_dataframe(prepared, params)
    assert pd.isna(wide.loc["Cr2O3", "s1"])


def test_mc_to_csv_rmelts_rejects_renormalization_mode(tmp_path):
    mc_df = pd.DataFrame(
        [
            {
                "SiO2": 75.0,
                "TiO2": 0.05,
                "Al2O3": 13.0,
                "FeO": 0.49,
                "Fe2O3": 0.18,
                "MnO": 0.03,
                "MgO": 0.08,
                "CaO": 0.6,
                "Na2O": 4.2,
                "K2O": 4.0,
                "P2O5": 0.01,
                "H2O": 2.0,
            }
        ]
    )
    mc_path = tmp_path / "mc.csv"
    mc_df.to_csv(mc_path, index=False)
    with pytest.raises(ValueError, match="Composition preservation is enforced"):
        rmp.MC_to_csv_rMELTS(mc_path, output_dir=tmp_path, dataset_name="x", validate_only=False)


def test_mc_to_csv_rmelts_fe3fet_bounds(tmp_path):
    mc_df = pd.DataFrame([{"SiO2": 75, "TiO2": 0.05, "Al2O3": 13, "FeOt": 1, "MnO": 0.03, "MgO": 0.1, "CaO": 0.5, "Na2O": 4, "K2O": 4, "P2O5": 0.01, "H2O": 2}])
    mc_path = tmp_path / "mc.csv"
    mc_df.to_csv(mc_path, index=False)
    with pytest.raises(ValueError):
        rmp.MC_to_csv_rMELTS(mc_path, output_dir=tmp_path, dataset_name="x", fe3fet=1.2)


def test_build_melts_input_wide_dataframe_row_order_and_labels():
    prepared_df = _minimal_prepared_df()
    params = rmp.MELTSRunParams(
        T1=1100, T2=700, dT=1, P1=400, P2=10, dP=10,
        fO2_constraint="buffered", fO2_buffer="NNO", fO2_offset=0.0
    )
    wide = rmp._build_melts_input_wide_dataframe(prepared_df, params)
    assert list(wide.index[:15]) == rmp.MELTS_REQUIRED_15_ROWS
    assert "ΔT" in wide.index
    assert "ΔP" in wide.index
    assert wide.loc["Model", "S1"] == "rhyolite-MELTS_v1.0.x"
    assert float(wide.loc["T1", "S1"]) == pytest.approx(1100.0)


def test_build_melts_input_wide_dataframe_visual_print_full_submission():
    prepared_df = _minimal_prepared_df()
    params = rmp.MELTSRunParams(
        T1=1100, T2=700, dT=1, P1=400, P2=10, dP=10,
        fO2_constraint="buffered", fO2_buffer="NNO", fO2_offset=0.0
    )

    wide = rmp._build_melts_input_wide_dataframe(prepared_df, params)

    # Visual inspection aid: print the exact wide-format RMELTS batch submission matrix.
    print("\n=== RMELTS BATCH SUBMISSION (FAKE SAMPLE) : START ===")
    with pd.option_context("display.max_rows", None, "display.max_columns", None, "display.width", 200):
        print(wide.to_string())
    print("=== RMELTS BATCH SUBMISSION (FAKE SAMPLE) : END ===\n")

    assert list(wide.index[:15]) == rmp.MELTS_REQUIRED_15_ROWS
    for key in [
        "Model",
        "Calculation",
        "T1",
        "ΔT",
        "P1",
        "ΔP",
        "fO2 constraint",
        "fO2 buffer",
        "fO2 offset",
    ]:
        assert key in wide.index

    assert "S1" in wide.columns
    assert float(wide.loc["SiO2", "S1"]) == pytest.approx(75.0)
    assert float(wide.loc["H2O", "S1"]) == pytest.approx(2.0)
    assert wide.loc["Model", "S1"] == "rhyolite-MELTS_v1.0.x"
    assert wide.loc["Calculation", "S1"] == "QF_P_Calc"
    assert float(wide.loc["T1", "S1"]) == pytest.approx(1100.0)
    assert float(wide.loc["P1", "S1"]) == pytest.approx(400.0)
    assert wide.loc["fO2 constraint", "S1"] == "buffered"
    assert wide.loc["fO2 buffer", "S1"] == "NNO"
    assert float(wide.loc["fO2 offset", "S1"]) == pytest.approx(0.0)


def test_rmelts_run_import_backend_writes_manifest_and_uses_run_dir(tmp_path, monkeypatch):
    prepared_df = _minimal_prepared_df()
    prepared_path = tmp_path / "prepared.csv"
    prepared_df.to_csv(prepared_path, index=False)

    class FakeHelperModule:
        @staticmethod
        def parallel_melts_main_loop(csv_path, max_composition_workers, max_pressure_workers, verbose):
            # Should run inside the run directory due to _pushd in _run_helper_import_backend.
            cwd = Path.cwd()
            outdir = cwd / "parallel-results"
            outdir.mkdir(exist_ok=True)
            fname = outdir / "Parallel_MELTS_S1_dP10_01-01_1cores.xlsx"
            fname.write_bytes(b"fake")
            return [
                {
                    "label": "S1",
                    "filename": str(fname.relative_to(cwd)),
                    "num_pressure_steps": 3,
                    "num_data_points": 10,
                    "P_QF": 123.4,
                    "P_Q2F": 150.0,
                    "total_time": 1.5,
                    "calc_time": 1.2,
                }
            ]

    monkeypatch.setattr(rmp, "_load_helper_module", lambda helper_dir, **kwargs: FakeHelperModule)

    result = rmp.rMELTS_run(
        prepared_path,
        output_dir=tmp_path,
        dataset_name="batch",
        T1=1100,
        T2=700,
        dT=1,
        P1=400,
        P2=10,
        dP=10,
        fO2_constraint="buffered",
        fO2_buffer="NNO",
        fO2_offset=0.0,
        helper_dir=tmp_path,  # not used because loader is monkeypatched
        backend="import",
        max_composition_workers=1,
        max_pressure_workers=1,
        verbose=False,
        pressure_residual_threshold_C=10.0,
    )

    manifest = pd.read_csv(result.manifest_csv_path)
    assert manifest.loc[0, "status"] == "success"
    assert Path(manifest.loc[0, "excel_path"]).exists()
    assert Path(result.run_dir).is_dir()
    assert Path(result.melts_input_csv_path).exists()
    assert result.summary["num_success"] == 1
    for col in [
        "melts_calc_time_s",
        "pressure_calc_time_s",
        "workbook_build_time_s",
        "runtime_log_version",
        "pressure_residual_threshold_C",
    ]:
        assert col in manifest.columns
    assert manifest.loc[0, "melts_calc_time_s"] == pytest.approx(1.2)
    assert manifest.loc[0, "runtime_log_version"] == "rmp_runtime_v1"
    assert manifest.loc[0, "pressure_residual_threshold_C"] == pytest.approx(10.0)


def _write_test_workbook(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws0 = wb.active
    ws0.title = "system"
    headers = ["Index", "T (C)", "P (MPa)", "G (kJ)", "H (kJ)", "S (J/K)", "Cp (J/K)", "V (cm3)", "mass (g)", "rho (g/cm3)", "Outcome"]
    ws0.append(headers)
    ws0.append([1, 800, 150, -1000, -900, 100, 50, 10, 100, 2.5, "Success"])

    ws1 = wb.create_sheet("liquid")
    ws1.append(headers + ["SiO2 (wt%)"])
    ws1.append([1, 800, 150, -500, -450, 80, 40, 9, 90, 2.3, "Success", 74.5])

    ws2 = wb.create_sheet("quartz")
    ws2.append(headers)
    ws2.append([1, 800, 150, -300, -280, 20, 10, 5, 10, 2.6, "Success"])

    wb.save(path)


def test_geobarometry_basis_extracts_selected_phases_and_reports_missing(tmp_path):
    wb_path = tmp_path / "sample.xlsx"
    _write_test_workbook(wb_path)

    manifest = pd.DataFrame(
        [
            {
                "dataset_name": "d",
                "sample_label": "S1",
                "status": "success",
                "excel_path": str(wb_path),
                "helper_filename": "parallel-results/sample.xlsx",
                "num_pressure_steps": 1,
                "num_data_points": 1,
                "P_QF": 100,
                "P_Q2F": 120,
                "total_time_s": 1,
                "calc_time_s": 1,
                "error": "",
                "run_dir": str(tmp_path),
                "melts_input_csv_path": str(tmp_path / "melts_input.csv"),
            }
        ]
    )
    manifest_path = tmp_path / "manifest.csv"
    manifest.to_csv(manifest_path, index=False)

    basis = rmp.rMELTS_geobarometry_basis(
        manifest_path,
        phases=["Quartz", "Kspar"],
        include_system=True,
        include_liquid=True,
        missing_phase_policy="skip_point",
    )

    assert "quartz" in basis.phase_tables
    assert "system" in basis.phase_tables
    assert "liquid" in basis.phase_tables
    assert not basis.phase_tables["quartz"].empty
    assert "G_kJ" in basis.phase_tables["quartz"].columns
    missing = basis.extraction_report["missing_phases"]
    assert any(row["phase"] == "kspar" for row in missing)
    avail = basis.phase_availability_table
    assert ((avail["requested_phase"] == "kspar") & (avail["present"] == False)).any()


def test_geobarometry_basis_fill_nan_missing_phase_policy(tmp_path):
    wb_path = tmp_path / "sample.xlsx"
    _write_test_workbook(wb_path)
    manifest_path = tmp_path / "manifest.csv"
    pd.DataFrame(
        [
            {
                "dataset_name": "d",
                "sample_label": "S1",
                "status": "success",
                "excel_path": str(wb_path),
                "helper_filename": "parallel-results/sample.xlsx",
                "num_pressure_steps": 1,
                "num_data_points": 1,
                "P_QF": 100,
                "P_Q2F": 120,
                "total_time_s": 1,
                "calc_time_s": 1,
                "error": "",
                "run_dir": str(tmp_path),
                "melts_input_csv_path": str(tmp_path / "melts_input.csv"),
            }
        ]
    ).to_csv(manifest_path, index=False)

    basis = rmp.rMELTS_geobarometry_basis(
        manifest_path,
        phases=["missing_phase"],
        include_system=False,
        include_liquid=False,
        missing_phase_policy="fill_nan",
    )
    assert "missing_phase" in basis.phase_tables
    df = basis.phase_tables["missing_phase"]
    assert len(df) == 1
    assert df.loc[0, "sample_label"] == "S1"
    assert pd.isna(df.loc[0, "T_C"])


def test_normalize_aplite_xrf_to_rowwise_mc_transposes_and_preserves_values(tmp_path):
    xrf = pd.DataFrame(
        [
            {"Sample Name": "SiO2", "KCP109B": 76.36, "KCP109C": 75.82},
            {"Sample Name": "TiO2", "KCP109B": 0.069, "KCP109C": 0.082},
            {"Sample Name": "Al2O3", "KCP109B": 12.65, "KCP109C": 12.95},
            {"Sample Name": "FeO*", "KCP109B": 0.52, "KCP109C": 0.66},
            {"Sample Name": "MnO", "KCP109B": 0.024, "KCP109C": 0.031},
            {"Sample Name": "MgO", "KCP109B": 0.05, "KCP109C": 0.09},
            {"Sample Name": "CaO", "KCP109B": 0.54, "KCP109C": 0.70},
            {"Sample Name": "Na2O", "KCP109B": 3.97, "KCP109C": 4.37},
            {"Sample Name": "K2O", "KCP109B": 4.61, "KCP109C": 4.01},
            {"Sample Name": "P2O5", "KCP109B": 0.006, "KCP109C": 0.009},
        ]
    )
    xrf_path = tmp_path / "Aplites_HAL_XRF_noUncertainty.csv"
    xrf.to_csv(xrf_path, index=False)

    result = rmp.normalize_aplite_xrf_to_rowwise_mc(
        xrf_path,
        output_dir=tmp_path,
        dataset_name="aplites",
        fixed_h2o_wt=13.0,
        target_samples=["KCP109B"],
    )
    all_df = pd.read_csv(result.normalized_all_rowwise_csv_path)
    one_df = pd.read_csv(result.normalized_target_rowwise_csv_path)

    assert result.num_samples == 2
    assert set(["SampleID", "FeOt", "H2O", "dry_total_original", "dry_total_preserved"]).issubset(all_df.columns)
    assert one_df["SampleID"].tolist() == ["KCP109B"]
    assert one_df.iloc[0]["H2O"] == pytest.approx(13.0)
    dry_cols = ["SiO2", "TiO2", "Al2O3", "FeOt", "MnO", "MgO", "CaO", "Na2O", "K2O", "P2O5"]
    # Values are preserved exactly (no in-function renormalization).
    assert one_df.iloc[0]["SiO2"] == pytest.approx(76.36)
    assert one_df.iloc[0]["FeOt"] == pytest.approx(0.52)
    assert one_df.iloc[0][dry_cols].sum() == pytest.approx(one_df.iloc[0]["dry_total_original"], abs=1e-12)
    assert one_df.iloc[0]["dry_total_preserved"] == pytest.approx(one_df.iloc[0]["dry_total_original"], abs=1e-12)


def test_pressure_analysis_parser_accepts_liam_and_pipeline_labels():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pressure Analysis"
    ws.append(["Phase System", "Estimated Pressure (MPa)", "Minimum Residual (°C)", "a (quadratic)", "b (linear)", "c (constant)"])
    ws.append(["2-Phase", 325.0, 2.5, 0.01, -6.5, 1000.0])
    ws.append(["3-Phase", 300.0, 4.0, 0.02, -12.0, 1800.0])
    parsed = rmp._compute_pressure_analysis_from_workbook(wb, prefer_existing_sheet=True)
    assert parsed["P_2phase_qtz_fsp_MPa"] == pytest.approx(325.0)
    assert parsed["P_3phase_qtz_fsp_fsp1_MPa"] == pytest.approx(300.0)

    ws["A2"] = "2-Phase (Qtz+Fsp)"
    ws["A3"] = "3-Phase (Qtz+Fsp+Fsp1)"
    parsed2 = rmp._compute_pressure_analysis_from_workbook(wb, prefer_existing_sheet=True)
    assert parsed2["P_2phase_qtz_fsp_MPa"] == pytest.approx(325.0)
    assert parsed2["P_3phase_qtz_fsp_fsp1_MPa"] == pytest.approx(300.0)


def test_first_appearance_extractor_and_staged_window_selector():
    phase_tables = {
        "quartz": pd.DataFrame(
            [
                {"P_MPa": 400, "T_C": 760, "Mass_g": 0.0},
                {"P_MPa": 400, "T_C": 755, "Mass_g": 1.0},
                {"P_MPa": 375, "T_C": 750, "Mass_g": 2.0},
            ]
        ),
        "feldspar": pd.DataFrame(
            [
                {"P_MPa": 400, "T_C": 752, "Mass_g": 0.2},
                {"P_MPa": 375, "T_C": 748, "Mass_g": 0.5},
            ]
        ),
        "feldspar_1": pd.DataFrame(
            [
                {"P_MPa": 375, "T_C": 742, "Mass_g": 0.1},
                {"P_MPa": 350, "T_C": None, "Mass_g": 0.1},
            ]
        ),
    }
    first_df = rmp._extract_first_appearance_table_from_phase_tables(phase_tables)
    assert not first_df.empty
    qz_400 = first_df[(first_df["phase"] == "quartz") & (first_df["P_MPa"] == 400)]
    # zero-mass row at 760 should be ignored; first valid appearance is 755
    assert qz_400.iloc[0]["T_first_appearance_C"] == pytest.approx(755)

    window = rmp._choose_staged_temperature_window(
        first_df,
        full_T1=1100,
        full_T2=600,
        margin_high_C=30,
        margin_low_C=80,
    )
    assert window["reason"] == "from_first_appearance"
    assert window["T1_fine"] <= 1100
    assert window["T2_fine"] >= 600
    assert window["T1_fine"] > window["T2_fine"]

    fallback = rmp._choose_staged_temperature_window(
        pd.DataFrame(columns=["phase", "P_MPa", "T_first_appearance_C"]),
        full_T1=1100,
        full_T2=600,
        margin_high_C=30,
        margin_low_C=80,
    )
    assert fallback["reason"] == "no_phase_appearance_rows"
    assert fallback["T1_fine"] == pytest.approx(1100)
    assert fallback["T2_fine"] == pytest.approx(600)


def test_load_helper_module_uses_stable_name_and_sys_path(tmp_path):
    helper_path = tmp_path / "MeltsHelperFunctions.py"
    helper_path.write_text("VALUE = 1\n", encoding="utf-8")
    old_pythonpath = os.environ.get("PYTHONPATH")
    observed_pythonpath = None
    try:
        if "PYTHONPATH" in os.environ:
            del os.environ["PYTHONPATH"]
        module = rmp._load_helper_module(tmp_path, run_dir=tmp_path)
        observed_pythonpath = os.environ.get("PYTHONPATH", "")
    finally:
        if old_pythonpath is None:
            os.environ.pop("PYTHONPATH", None)
        else:
            os.environ["PYTHONPATH"] = old_pythonpath
    assert module.__name__ == "MeltsHelperFunctions"
    assert str(tmp_path) in rmp.sys.path
    assert observed_pythonpath is not None
    assert str(tmp_path) in observed_pythonpath


def test_load_helper_module_keeps_run_local_helper_ahead_of_original_for_spawns(tmp_path):
    helper_dir = tmp_path / "helper"
    run_dir = tmp_path / "run"
    helper_dir.mkdir()
    helper_path = helper_dir / "MeltsHelperFunctions.py"
    helper_path.write_text("VALUE = 1\n", encoding="utf-8")

    old_pythonpath = os.environ.get("PYTHONPATH")
    try:
        if "PYTHONPATH" in os.environ:
            del os.environ["PYTHONPATH"]
        module = rmp._load_helper_module(helper_dir, run_dir=run_dir, patch_pressure_calc=False)
        run_dir_str = str(run_dir.resolve())
        helper_dir_str = str(helper_dir.resolve())
        assert module.__name__ == "MeltsHelperFunctions"
        assert run_dir_str in rmp.sys.path
        assert helper_dir_str in rmp.sys.path
        assert rmp.sys.path.index(run_dir_str) < rmp.sys.path.index(helper_dir_str)
        assert os.environ.get("PYTHONPATH", "").split(os.pathsep)[0] == run_dir_str
    finally:
        if old_pythonpath is None:
            os.environ.pop("PYTHONPATH", None)
        else:
            os.environ["PYTHONPATH"] = old_pythonpath


def test_load_helper_module_supports_liam_clone_implementation(tmp_path, monkeypatch):
    helper_dir = tmp_path / "helper"
    run_dir = tmp_path / "run"
    clone_dir = tmp_path / "clone"
    helper_dir.mkdir()
    clone_dir.mkdir()
    (helper_dir / "MeltsHelperFunctions.py").write_text("VALUE = 1\n", encoding="utf-8")
    clone_path = clone_dir / "rmelts_liam_clone_helper.py"
    clone_path.write_text("VALUE = 42\n", encoding="utf-8")

    monkeypatch.setattr(rmp, "_repo_local_liam_clone_helper_path", lambda: clone_path)

    module = rmp._load_helper_module(
        helper_dir,
        run_dir=run_dir,
        helper_implementation="liam_clone",
        patch_pressure_calc=False,
    )
    assert module.__name__ == "MeltsHelperFunctions"
    assert getattr(module, "VALUE") == 42
    assert (run_dir / "MeltsHelperFunctions.py").exists()


def test_resolve_helper_implementation_spec_supports_liam_clone_min_safe(tmp_path, monkeypatch):
    clone_path = tmp_path / "rmelts_liam_clone_helper.py"
    clone_path.write_text("VALUE = 42\n", encoding="utf-8")
    monkeypatch.setattr(rmp, "_repo_local_liam_clone_helper_path", lambda: clone_path)

    spec = rmp._resolve_helper_implementation_spec("liam_clone_min_safe")
    assert spec.name == "liam_clone_min_safe"
    assert spec.source_path == clone_path
    assert spec.patch_profile == "profile_h_production_liquidus_raw_reset"


def test_resolve_helper_implementation_spec_supports_liam_clone_guarded_main(tmp_path, monkeypatch):
    clone_path = tmp_path / "rmelts_liam_clone_helper.py"
    clone_path.write_text("VALUE = 42\n", encoding="utf-8")
    monkeypatch.setattr(rmp, "_repo_local_liam_clone_helper_path", lambda: clone_path)

    spec = rmp._resolve_helper_implementation_spec("liam_clone_guarded_main")
    assert spec.name == "liam_clone_guarded_main"
    assert spec.source_path == clone_path
    assert spec.patch_profile == "profile_i_prototype_mainloop_bounded_reset"


def test_patch_helper_source_text_for_spawn_safety_increments_timeout_loop_counter():
    src = (
        "def f():\n"
        "    if True:\n"
        "                current_T = current_T+25  # continue with higher temperature\n"
        "                continue\n"
        "                state = equil.execute(temp + 273.15, pressure * 10, bulk_comp=blk_cmp, con_deltaNNO=fO2_offset, debug=0)\n"
        "\n"
        "                t1 = time.time()\n"
        "                calc_time += (t1 - t0)\n"
        "        temp = find_wet_liquidus(equil, T1, T2, pressure, 50, blk_cmp, fO2_offset, verbose)\n"
    )
    out = rmp._patch_helper_source_text_for_spawn_safety(src, patch_profile="profile_e_current_full_patch")
    assert "i = i + 1" in out
    assert "min(_rmelts_T_upper_bound, current_T+25)" in out
    assert "safe_equilibrium_execute" in out
    assert "timeout in execute" in out
    assert "skip wet-liquidus pre-search" in out


def test_patch_helper_source_text_for_spawn_safety_production_default_resets_liquidus_solver():
    src = (
        "def find_wet_liquidus(equil, T1, T2, P, n, composition, fO2_offset, verbose=False):\n"
        "    current_T = round((T1+T2)/2)\n"
        "    i = 0\n"
        "    max_iterations = 50\n"
        "    timeout = 10\n"
        "    dbg = 0\n"
        "    try:\n"
        "        state = safe_equilibrium_execute(equil, current_T+273.15, P*10, timeout=timeout, \\n"
        "                                       bulk_comp=composition, con_deltaNNO=fO2_offset, debug=dbg)\n"
        "        if state is None:\n"
        "            logging.error(f\"Initial equilibrium calculation failed or timed out at T={current_T}°C\")\n"
        "            return T1  # Return fallback temperature\n"
        "    except Exception as e:\n"
        "        logging.error(f\"Initial calculation error: {e}\")\n"
        "        return T1\n"
        "    while (T1 > T2) and i < 50:\n"
        "        try:\n"
        "            state = safe_equilibrium_execute(equil, current_T+273.15, P*10, timeout=timeout, \\n"
        "                                        state=state, con_deltaNNO=fO2_offset, debug=dbg)\n"
        "            if state is None:\n"
        "                logging.error(f\"Equilibrium calculation failed or timed out at T={current_T}°C\")\n"
        "                current_T = current_T+25  # continue with higher temperature\n"
        "                continue\n"
        "            if(len(phases) > min_phases):\n"
        "                state = safe_equilibrium_execute(equil, current_T+1+273.15, P*10, state=state,timeout=timeout, \\n"
        "                                         con_deltaNNO=fO2_offset, debug=dbg)\n"
        "                if state is None:\n"
        "                    logging.error(f\"Equilibrium calculation failed or timed out at T={current_T}°C\")\n"
        "                    return T1  # Return fallback temperature\n"
        "        except Exception as e:\n"
        "            print(e)\n"
        "            return T1\n"
    )
    out = rmp._patch_helper_source_text_for_spawn_safety(src)
    assert "# rMELTS pipeline bootstrap: prefer this run-local helper copy in spawned workers" in out
    assert "_rmelts_helper_dir" in out
    assert "_rmelts_reset_equil" in out
    assert "_rmelts_liquidus_execute" in out
    assert "_rmelts_liquidus_bulk_comp = composition" in out
    assert "equilibrate.Equilibrate(equil.element_list, equil.phase_list)" in out
    assert "state.set_phase_comp(omni_phase, _rmelts_liquidus_bulk_comp, input_as_elements=True)" in out
    assert "return equil_obj.execute" in out
    assert "Liquidus raw execute failed" in out
    assert "LIQUIDUS_RESET_BEGIN" in out
    assert "LIQUIDUS_RESET_RESEED_OK" in out
    assert "LIQUIDUS_RESET_RESEED_FAIL" in out
    assert "LIQUIDUS_RETRY_WITH_CANONICAL_BULK" in out
    assert "LIQUIDUS_RETRY_WITH_CALL_BULK" in out
    assert "LIQUIDUS_RETRY_FAILED" in out
    assert "bulk_comp=_rmelts_retry_bulk_comp" in out
    assert "_rmelts_T_upper_bound = T1" in out
    assert "_rmelts_bounds_invalid" in out
    assert "min(_rmelts_T_upper_bound, current_T+25)" in out
    assert "skip wet-liquidus pre-search" not in out
    # Production default should not patch run_single_pressure_step main execute timeout wrapper.
    assert "timeout in execute" not in out


def test_patch_helper_liquidus_retry_after_reset_never_uses_none_bulk_comp():
    src = (
        "def find_wet_liquidus(equil, T1, T2, P, n, composition, fO2_offset, verbose=False):\n"
        "    current_T = round((T1+T2)/2)\n"
        "    i = 0\n"
        "    max_iterations = 50\n"
        "    timeout = 10\n"
        "    dbg = 0\n"
        "    state = safe_equilibrium_execute(equil, current_T+273.15, P*10, timeout=timeout, \\n"
        "                                       bulk_comp=composition, con_deltaNNO=fO2_offset, debug=dbg)\n"
    )
    out = rmp._patch_helper_source_text_for_spawn_safety(src, patch_profile="profile_h_production_liquidus_raw_reset")
    # Regression guard: retry path must substitute canonical bulk composition if call-site bulk_comp is None.
    assert "_rmelts_retry_bulk_comp = bulk_comp" in out
    assert "_rmelts_retry_bulk_comp = _rmelts_liquidus_bulk_comp" in out
    assert "bulk_comp=_rmelts_retry_bulk_comp" in out
    assert "bulk_comp=bulk_comp" not in out.split("return equil.execute", 1)[1]


def test_patch_helper_fresh_liquidus_profiles_inject_iteration_and_branch_resets():
    src = (
        "def find_wet_liquidus(equil, T1, T2, P, n, composition, fO2_offset, verbose=False):\n"
        "    current_T = round((T1+T2)/2)\n"
        "    i = 0\n"
        "    timeout = 10\n"
        "    dbg = 0\n"
        "    while (T1 > T2) and i < 50:\n"
        "        try:\n"
        "            state = safe_equilibrium_execute(equil, current_T+273.15, P*10, timeout=timeout, \\n"
        "                                        state=state, con_deltaNNO=fO2_offset, debug=dbg)\n"
        "            if state is None:\n"
        "                logging.error(f\"Equilibrium calculation failed or timed out at T={current_T}°C\")\n"
        "                current_T = current_T+25  # continue with higher temperature\n"
        "                continue\n"
        "        except Exception as e:\n"
        "            print(e)\n"
        "            return T1\n"
        "def run_single_pressure_step(args):\n"
        "    temp = find_wet_liquidus(equil, T1, T2, pressure, 50, blk_cmp, fO2_offset, verbose)\n"
        "    for step in range(N_runs + 1):\n"
        "        try:\n"
        "                state = equil.execute(temp + 273.15, pressure * 10, bulk_comp=blk_cmp, con_deltaNNO=fO2_offset, debug=0)\n"
        "\n"
        "                t1 = time.time()\n"
        "                calc_time += (t1 - t0)\n"
        "        except Exception as e:\n"
        "            print(f\"Error at T={temp}, P={pressure}: {e}\")\n"
        "            pass\n"
    )
    out_j = rmp._patch_helper_source_text_for_spawn_safety(
        src, patch_profile="profile_j_liquidus_fresh_iteration_bounded_main"
    )
    out_k = rmp._patch_helper_source_text_for_spawn_safety(
        src, patch_profile="profile_k_liquidus_fresh_branch_bounded_main"
    )
    assert "LIQUIDUS_FRESH_ITERATION_RESET" in out_j
    assert "LIQUIDUS_FRESH_BRANCH_CANONICAL_BULK" not in out_j
    assert "LIQUIDUS_FRESH_BRANCH_CANONICAL_BULK" in out_k
    assert "LIQUIDUS_FRESH_BRANCH_CALL_BULK" in out_k
    assert "_rmelts_liquidus_bulk_comp = composition" in out_j
    assert "_rmelts_liquidus_bulk_comp = composition" in out_k


def test_patch_helper_profile_l_reraises_non_timeout_safe_execute_errors():
    src = (
        "def safe_equilibrium_execute(equil, T, P, timeout=30, **kwargs):\n"
        "    try:\n"
        "        state = protected_execute(T, P, **kwargs)\n"
        "        if state is None:\n"
        "            raise RuntimeError(\"Equilibrium calculation returned None\")\n"
        "        return state\n"
        "    except TimeoutError as e:\n"
        "        logging.warning(f\"Equilibrium calculation timed out: T={T:.1f}K, P={P:.1f}bar - {str(e)}\")\n"
        "        return None\n"
        "        \n"
        "    except Exception as e:\n"
        "        logging.warning(f\"Equilibrium calculation failed: T={T:.1f}K, P={P:.1f}bar - {str(e)}\")\n"
        "        return None\n"
    )
    out = rmp._patch_helper_source_text_for_spawn_safety(
        src, patch_profile="profile_l_main_timeout_wrapper_reraise_non_timeout"
    )
    assert "except TimeoutError as e:" in out
    assert "return None" in out
    assert "except Exception as e:" in out
    # Non-timeout safe-equilibrium errors should propagate to the main-loop exception path.
    assert "logging.warning(f\"Equilibrium calculation failed" in out
    assert "raise" in out.split("except Exception as e:", 1)[1]


def test_patch_helper_source_uses_profile_main_execute_timeout():
    src = (
        "def safe_equilibrium_execute(equil, T, P, timeout=30, **kwargs):\n"
        "    pass\n\n"
        "def find_wet_liquidus(equil, T1, T2, P, step, composition, fO2_offset, dbg):\n"
        "    timeout = 10\n"
        "    while (T1 > T2) and i < 50:\n"
        "        try:\n"
        "            pass\n"
        "        except Exception:\n"
        "            pass\n"
        "def run_single_pressure_step(args):\n"
        "    temp = find_wet_liquidus(equil, T1, T2, pressure, 50, blk_cmp, fO2_offset, verbose)\n"
        "    for step in range(N_runs + 1):\n"
        "        try:\n"
        "                state = equil.execute(temp + 273.15, pressure * 10, bulk_comp=blk_cmp, con_deltaNNO=fO2_offset, debug=0)\n"
        "\n"
        "                t1 = time.time()\n"
        "                calc_time += (t1 - t0)\n"
        "        except Exception as e:\n"
        "            print(f\"Error at T={temp}, P={pressure}: {e}\")\n"
        "            pass\n"
    )
    custom = replace(
        rmp._HELPER_PATCH_PROFILES["profile_k_liquidus_fresh_branch_bounded_main"],
        name="profile_k_timeout_42_test",
        main_execute_timeout_s=42.0,
    )
    out = rmp._patch_helper_source_text_for_spawn_safety(src, patch_profile=custom)
    assert "timeout=42.0" in out


def test_safe_float_handles_numeric_and_non_numeric_values():
    np = pytest.importorskip("numpy")
    assert rmp._safe_float(None) is None
    assert rmp._safe_float("Fit Failed") is None
    assert rmp._safe_float(1) == pytest.approx(1.0)
    assert rmp._safe_float("2.5") == pytest.approx(2.5)
    assert rmp._safe_float(float("nan")) is None
    assert rmp._safe_float(np.float64(3.2)) == pytest.approx(3.2)


def test_pressure_analysis_parser_handles_fit_failed_cells():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pressure Analysis"
    ws.append(
        [
            "Phase System",
            "Estimated Pressure (MPa)",
            "Minimum Residual (°C)",
            "a (quadratic)",
            "b (linear)",
            "c (constant)",
        ]
    )
    ws.append(["2-Phase", "Fit Failed", 8.2, None, None, None])
    ws.append(["3-Phase", 301.0, 4.1, 0.02, -12.0, 1800.0])
    parsed = rmp._read_pressure_analysis_sheet_from_workbook(wb)
    assert parsed["P_2phase_qtz_fsp_MPa"] is None
    assert parsed["Rmin_2phase_qtz_fsp_C"] == pytest.approx(8.2)
    assert parsed["P_3phase_qtz_fsp_fsp1_MPa"] == pytest.approx(301.0)


def test_pipeline_fallback_pressure_threshold_acceptance():
    wb = openpyxl.Workbook()
    ws_qz = wb.active
    ws_qz.title = "quartz"
    headers = ["Index", "T (C)", "P (MPa)", "mass (g)"]
    ws_qz.append(headers)
    ws_qz.append([1, 800, 400, 1.0])
    ws_qz.append([2, 790, 300, 1.0])
    ws_qz.append([3, 780, 200, 1.0])

    ws_f = wb.create_sheet("feldspar")
    ws_f.append(headers)
    ws_f.append([1, 788, 400, 1.0])  # |800-788| = 12
    ws_f.append([2, 783, 300, 1.0])  # |790-783| = 7
    ws_f.append([3, 768, 200, 1.0])  # |780-768| = 12

    ws_f1 = wb.create_sheet("feldspar_1")
    ws_f1.append(headers)
    ws_f1.append([1, 770, 400, 1.0])  # 3-phase spread = 30
    ws_f1.append([2, 771, 300, 1.0])  # 3-phase spread = 19
    ws_f1.append([3, 760, 200, 1.0])  # 3-phase spread = 20

    parsed5 = rmp._compute_pressure_analysis_from_workbook(wb, prefer_existing_sheet=False, residual_threshold=5.0)
    parsed10 = rmp._compute_pressure_analysis_from_workbook(wb, prefer_existing_sheet=False, residual_threshold=10.0)

    assert parsed5["P_2phase_qtz_fsp_MPa"] is None
    assert parsed10["P_2phase_qtz_fsp_MPa"] is not None
    assert parsed10["Rmin_2phase_qtz_fsp_C"] is not None
    assert parsed10["Rmin_2phase_qtz_fsp_C"] <= 10.0
    assert parsed10["P_3phase_qtz_fsp_fsp1_MPa"] is None


def test_pipeline_fallback_uses_raw_residual_threshold_before_fit(monkeypatch):
    wb = openpyxl.Workbook()
    ws_qz = wb.active
    ws_qz.title = "quartz"
    headers = ["Index", "T (C)", "P (MPa)", "mass (g)"]
    ws_qz.append(headers)
    ws_qz.append([1, 800, 400, 1.0])
    ws_qz.append([2, 790, 300, 1.0])
    ws_qz.append([3, 780, 200, 1.0])

    ws_f = wb.create_sheet("feldspar")
    ws_f.append(headers)
    ws_f.append([1, 790, 400, 1.0])  # res2 = 10
    ws_f.append([2, 780, 300, 1.0])  # res2 = 10
    ws_f.append([3, 770, 200, 1.0])  # res2 = 10

    ws_f1 = wb.create_sheet("feldspar_1")
    ws_f1.append(headers)
    ws_f1.append([1, 790, 400, 1.0])  # res3 = 10
    ws_f1.append([2, 780, 300, 1.0])  # res3 = 10
    ws_f1.append([3, 770, 200, 1.0])  # res3 = 10

    # Force fitted minima above threshold to verify acceptance still uses raw minima.
    def fake_fit(pressures, residuals):
        return (250.0, 10.8, (0.01, -1.0, 100.0))

    monkeypatch.setattr(rmp, "_fit_pressure_residual_vertex", fake_fit)
    parsed = rmp._compute_pressure_analysis_from_workbook(wb, prefer_existing_sheet=False, residual_threshold=10.0)
    assert parsed["P_2phase_qtz_fsp_MPa"] == pytest.approx(250.0)
    assert parsed["P_3phase_qtz_fsp_fsp1_MPa"] == pytest.approx(250.0)
    # Fitted residual is still reported for diagnostics even when > threshold.
    assert parsed["Rmin_3phase_qtz_fsp_fsp1_C"] == pytest.approx(10.8)


def test_create_manifest_parses_runtime_fields_with_safe_float(tmp_path):
    run_dir = tmp_path / "run"
    run_dir.mkdir()
    manifest_df = rmp._create_manifest_from_results(
        dataset_name="demo",
        run_dir=run_dir,
        melts_input_csv_path=run_dir / "melts_input.csv",
        expected_labels=["S1"],
        results=[
            {
                "label": "S1",
                "filename": "",
                "error": "mock error",
                "calc_time": "1.25",
                "total_time": "2.5",
                "melts_calc_time": "1.25",
                "pressure_calc_time": "0.5",
                "workbook_build_time": "0.75",
            }
        ],
        pressure_residual_threshold_C=10.0,
    )
    row = manifest_df.iloc[0]
    assert row["melts_calc_time_s"] == pytest.approx(1.25)
    assert row["pressure_calc_time_s"] == pytest.approx(0.5)
    assert row["workbook_build_time_s"] == pytest.approx(0.75)
    assert row["total_time_s"] == pytest.approx(2.5)
    assert row["pressure_residual_threshold_C"] == pytest.approx(10.0)


def test_rmelts_run_passes_pressure_threshold_to_import_backend(tmp_path, monkeypatch):
    prepared_df = _minimal_prepared_df()
    prepared_path = tmp_path / "prepared.csv"
    prepared_df.to_csv(prepared_path, index=False)

    captured: dict[str, object] = {}

    def fake_run_helper_import_backend(**kwargs):
        captured.update(kwargs)
        return [
            {
                "label": "S1",
                "filename": "",
                "error": "mock helper fail",
                "num_pressure_steps": None,
                "num_data_points": None,
                "calc_time": 1.0,
                "total_time": 1.5,
            }
        ]

    monkeypatch.setattr(rmp, "_run_helper_import_backend", fake_run_helper_import_backend)

    result = rmp.rMELTS_run(
        prepared_path,
        output_dir=tmp_path,
        dataset_name="threshold_pass",
        T1=1100,
        T2=700,
        dT=1,
        P1=400,
        P2=10,
        dP=10,
        fO2_constraint="buffered",
        fO2_buffer="NNO",
        fO2_offset=0.0,
        helper_dir=tmp_path,
        backend="import",
        max_composition_workers=1,
        max_pressure_workers=1,
        pressure_residual_threshold_C=10.0,
    )
    assert Path(result.manifest_csv_path).exists()
    assert captured["pressure_residual_threshold_C"] == pytest.approx(10.0)


def test_rmelts_run_records_cleanup_summary_in_metadata(tmp_path, monkeypatch):
    prepared_df = _minimal_prepared_df()
    prepared_path = tmp_path / "prepared.csv"
    prepared_df.to_csv(prepared_path, index=False)

    def fake_run_helper_import_backend(**kwargs):
        return [
            {
                "label": "S1",
                "filename": "",
                "error": "mock helper fail",
                "num_pressure_steps": None,
                "num_data_points": None,
                "calc_time": 1.0,
                "total_time": 1.5,
            }
        ]

    monkeypatch.setattr(rmp, "_run_helper_import_backend", fake_run_helper_import_backend)
    monkeypatch.setattr(
        rmp,
        "_snapshot_descendant_processes",
        lambda: {"parent_pid": 123, "psutil_available": True, "descendant_pids": [10], "warnings": []},
    )
    monkeypatch.setattr(
        rmp,
        "_cleanup_descendant_processes",
        lambda snapshot: {
            "cleanup_attempted": True,
            "cleanup_parent_pid": 123,
            "cleanup_psutil_available": True,
            "cleanup_descendants_before_count": 1,
            "cleanup_descendants_after_count": 3,
            "cleanup_descendants_found": 2,
            "cleanup_target_pids": [11, 12],
            "cleanup_terminated_count": 1,
            "cleanup_killed_count": 1,
            "cleanup_remaining_count": 0,
            "cleanup_duration_s": 0.25,
            "cleanup_warnings": [],
        },
    )

    result = rmp.rMELTS_run(
        prepared_path,
        output_dir=tmp_path,
        dataset_name="cleanup_meta",
        T1=1100,
        T2=700,
        dT=1,
        P1=400,
        P2=10,
        dP=10,
        fO2_constraint="buffered",
        fO2_buffer="NNO",
        fO2_offset=0.0,
        helper_dir=tmp_path,
        backend="import",
        max_composition_workers=1,
        max_pressure_workers=2,
        pressure_residual_threshold_C=10.0,
    )

    assert result.summary["process_cleanup"]["cleanup_descendants_found"] == 2
    metadata_path = Path(result.run_dir) / "run_metadata.json"
    metadata = json.loads(metadata_path.read_text(encoding="utf-8"))
    assert metadata["cleanup"]["cleanup_attempted"] is True
    assert metadata["cleanup"]["cleanup_terminated_count"] == 1
    assert metadata["cleanup"]["cleanup_killed_count"] == 1
    assert metadata["summary"]["helper_patch_profile"] == "profile_h_production_liquidus_raw_reset"


def test_rmelts_run_records_internal_helper_implementation_selector(tmp_path, monkeypatch):
    prepared_df = _minimal_prepared_df()
    prepared_path = tmp_path / "prepared.csv"
    prepared_df.to_csv(prepared_path, index=False)

    def fake_run_helper_import_backend(**kwargs):
        return [
            {
                "label": "S1",
                "filename": "",
                "error": "mock helper fail",
                "num_pressure_steps": None,
                "num_data_points": None,
                "calc_time": 1.0,
                "total_time": 1.5,
            }
        ]

    monkeypatch.setattr(rmp, "_run_helper_import_backend", fake_run_helper_import_backend)
    monkeypatch.setenv("RMELTS_INTERNAL_HELPER_IMPLEMENTATION", "patched_profile_k")

    result = rmp.rMELTS_run(
        prepared_path,
        output_dir=tmp_path,
        dataset_name="helper_impl_meta",
        T1=1100,
        T2=700,
        dT=1,
        P1=400,
        P2=10,
        dP=10,
        fO2_constraint="buffered",
        fO2_buffer="NNO",
        fO2_offset=0.0,
        helper_dir=tmp_path,
        backend="import",
        max_composition_workers=1,
        max_pressure_workers=1,
    )

    metadata_path = Path(result.run_dir) / "run_metadata.json"
    metadata = json.loads(metadata_path.read_text(encoding="utf-8"))
    assert metadata["summary"]["helper_implementation_requested"] == "patched_profile_k"
    assert metadata["summary"]["helper_implementation_resolved"] == "patched_profile_k"
    assert metadata["summary"]["helper_patch_profile"] == "profile_k_liquidus_fresh_branch_bounded_main"


def test_populate_workbook_deltaqfm_columns_uses_redox_buffer_and_existing_tabs(monkeypatch):
    wb = openpyxl.Workbook()
    init_ws = wb.active
    init_ws.title = "init_cond"
    init_ws.cell(row=9, column=4, value="fO2 value")
    init_ws.cell(row=9, column=5, value=0.5)
    init_ws.cell(row=10, column=4, value="fO2 buffer")
    init_ws.cell(row=10, column=5, value="nno")

    qz = wb.create_sheet("quartz")
    headers = ["Index", "T (C)", "P (MPa)", "deltaQFM", "mass (g)"]
    for col, h in enumerate(headers, 1):
        qz.cell(row=1, column=col, value=h)
    qz.cell(row=2, column=1, value=1)
    qz.cell(row=2, column=2, value=900.0)
    qz.cell(row=2, column=3, value=250.0)
    qz.cell(row=2, column=4, value=0.0)
    qz.cell(row=2, column=5, value=1.0)

    pa = wb.create_sheet("Pressure Analysis")
    pa.cell(row=1, column=1, value="Phase System")

    def fake_redox(T_K, P_bar, buffer_name):
        assert T_K == pytest.approx(1173.15)
        assert P_bar == pytest.approx(2500.0)
        if str(buffer_name).upper() == "NNO":
            return -11.0
        if str(buffer_name).upper() == "QFM":
            return -12.0
        raise AssertionError(f"Unexpected buffer {buffer_name}")

    monkeypatch.setattr(rmp, "_redox_buffer_logfO2", fake_redox)

    sheet_count_before = len(wb.sheetnames)
    summary = rmp._populate_workbook_deltaqfm_columns(wb)
    sheet_count_after = len(wb.sheetnames)

    assert sheet_count_after == sheet_count_before
    assert summary["deltaqfm_enriched"] is True
    assert summary["deltaqfm_sheets_updated"] == 1
    assert summary["deltaqfm_rows_updated"] == 1
    assert summary["deltaqfm_buffer"] == "NNO"
    assert qz.cell(row=2, column=4).value == pytest.approx(1.5)


def _build_minimal_melts_excel_template_workbook(path: Path) -> None:
    from openpyxl.workbook.defined_name import DefinedName

    wb = openpyxl.Workbook()
    ws_mc = wb.active
    ws_mc.title = "Multiple_Comp"
    ws_mc.cell(row=1, column=1, value=" ")
    ws_mc.cell(row=1, column=2, value="P_Calc")
    for i, oxide in enumerate(rmp.MELTS_OXIDE_ROWS, start=2):
        ws_mc.cell(row=i, column=1, value=oxide)

    # Row labels used by the writer on Multiple_Comp.
    mc_labels = {
        22: "Model",
        23: "Calculation",
        24: "T1",
        25: "T2",
        26: "ΔT",
        27: "T unit",
        28: "P1",
        29: "P2",
        30: "ΔP",
        31: "P unit",
        32: "fO2 offset",
        33: "fO2 buffer",
        34: "fO2 constraint",
        35: "Starting T",
        36: "Min liq content",
        37: "Fractionate",
        38: "Phase 1",
        39: "Phase 2",
        40: "Phase 3",
        41: "Formula",
        42: "ΔH",
        43: "ΔV",
        44: "ΔS",
    }
    for r, label in mc_labels.items():
        ws_mc.cell(row=r, column=1, value=label)

    ws_in = wb.create_sheet("Input")
    for i, oxide in enumerate(rmp.MELTS_OXIDE_ROWS, start=2):
        ws_in.cell(row=i, column=1, value=oxide)
    ws_in["D21"] = "Status placeholder"
    ws_in["E2"] = 0
    ws_in["E3"] = 0
    ws_in["E4"] = 0

    ws_seq = wb.create_sheet("Sequences")
    seq_labels = {
        2: "T1",
        3: "T2",
        4: "ΔT",
        6: "P1",
        7: "P2",
        8: "ΔP",
        10: "fO2",
        13: "Starting T",
        14: "Min liq content",
        15: "Fractionate",
        22: "Phase 1",
        23: "Phase 2",
        24: "Phase 3",
        25: "Formula",
        42: "ΔH",
        44: "ΔV",
        54: "ΔS",
    }
    for r, label in seq_labels.items():
        ws_seq.cell(row=r, column=1, value=label)

    wb.defined_names.add(DefinedName("Pressure", attr_text="Input!$E$2"))
    wb.defined_names.add(DefinedName("Temperature", attr_text="Input!$E$3"))
    wb.defined_names.add(DefinedName("log_fO2", attr_text="Input!$E$4"))

    wb.save(path)
    wb.close()


def test_write_composition_to_melts_excel_template_populates_multiple_comp_input_sequences(tmp_path):
    template_path = tmp_path / "melts_template.xlsm"
    output_path = tmp_path / "filled_template.xlsm"
    _build_minimal_melts_excel_template_workbook(template_path)

    comp = {
        "sample_label": "KCP109B-101525-v1",
        "SiO2": 76.36,
        "TiO2": 0.07,
        "Al2O3": 12.65,
        "FeO": 0.52,
        "MnO": 0.02,
        "MgO": 0.031,
        "CaO": 0.54,
        "Na2O": 3.97,
        "K2O": 4.61,
        "P2O5": 0.01,
        "H2O": 11.0,
    }
    params = rmp.MELTSRunParams(
        T1=900,
        T2=700,
        dT=1,
        P1=250,
        P2=150,
        dP=5,
        fO2_constraint="TRUE",
        fO2_buffer="nno",
        fO2_offset=0.0,
        model="rhyolite-MELTS_v1.0.x",
        calculation="QF_P_Calc",
    )

    result = rmp.write_composition_to_melts_excel_template(
        template_workbook_path=template_path,
        output_workbook_path=output_path,
        composition=comp,
        params=params,
        target_column_header="P_Calc",
        mirror_input_and_sequences=True,
        starting_T="wet liquidus",
        min_liq_content=10,
        phase_1="quartz",
        phase_2="feldspar1",
        phase_3="feldspar2",
        formula="any two phases",
    )

    assert result.output_workbook_path == str(output_path)
    assert result.target_column_header == "P_Calc"
    assert result.multiple_comp_oxide_order_matches_expected is True
    assert result.rows_written_multiple_comp > 0
    assert result.rows_written_input_sheet > 0
    assert result.rows_written_sequences > 0

    wb = openpyxl.load_workbook(output_path, keep_vba=True, data_only=False)
    try:
        mc = wb["Multiple_Comp"]
        # P_Calc is column 2 in the synthetic template.
        assert mc.cell(row=2, column=2).value == pytest.approx(76.36)  # SiO2
        assert mc.cell(row=5, column=2).value is None  # Fe2O3 intentionally blank
        assert mc.cell(row=7, column=2).value == pytest.approx(0.52)  # FeO
        assert mc.cell(row=16, column=2).value == pytest.approx(11.0)  # H2O
        assert mc.cell(row=23, column=2).value == "QF_P_Calc"
        assert mc.cell(row=24, column=2).value == pytest.approx(900.0)
        assert mc.cell(row=28, column=2).value == pytest.approx(250.0)
        assert mc.cell(row=33, column=2).value == "NNO"
        assert mc.cell(row=34, column=2).value is True

        ws_in = wb["Input"]
        assert ws_in.cell(row=2, column=2).value == pytest.approx(76.36)
        assert ws_in.cell(row=5, column=2).value is None  # Fe2O3 blank preserved
        assert ws_in.cell(row=7, column=2).value == pytest.approx(0.52)
        assert ws_in["E2"].value == pytest.approx(250.0)  # Pressure named cell mirror
        assert ws_in["E3"].value == pytest.approx(900.0)  # Temperature named cell mirror
        assert ws_in["E4"].value == pytest.approx(0.0)    # log_fO2 named cell mirror

        ws_seq = wb["Sequences"]
        assert ws_seq.cell(row=2, column=2).value == pytest.approx(900.0)
        assert ws_seq.cell(row=3, column=2).value == pytest.approx(700.0)
        assert ws_seq.cell(row=6, column=2).value == pytest.approx(250.0)
        assert ws_seq.cell(row=7, column=2).value == pytest.approx(150.0)
        assert ws_seq.cell(row=8, column=2).value == pytest.approx(5.0)
        assert ws_seq.cell(row=10, column=2).value == pytest.approx(0.0)
        assert ws_seq.cell(row=22, column=2).value == "quartz"
        assert ws_seq.cell(row=23, column=2).value == "feldspar1"
        assert ws_seq.cell(row=24, column=2).value == "feldspar2"
    finally:
        wb.close()
