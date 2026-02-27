from openpyxl import Workbook
import pandas as pd

from sci_helpers.melts_data_extractor import (
    extract_calls_long,
    extract_calls_wide,
    extract_workbook_tables,
    write_workbook_tables,
)


def _build_test_workbook(path):
    wb = Workbook()
    ws = wb.active
    ws.title = "init_cond"
    ws["A1"] = "SiO2"
    ws["B1"] = 76.2

    system = wb.create_sheet("system")
    system.append(["Index", "T (C)", "P (MPa)", "mass (g)", "Outcome"])
    system.append([1, 900, 200, 100.0, "Success"])

    liquid = wb.create_sheet("liquid")
    liquid.append(["Index", "T (C)", "P (MPa)", "mass (g)", "SiO2 (wt%)", "Outcome"])
    liquid.append([1, 900, 200, 80.0, 76.2, "Success"])

    quartz = wb.create_sheet("quartz")
    quartz.append(["Index", "T (C)", "P (MPa)", "mass (g)", "Outcome"])
    quartz.append([1, 900, 200, 2.0, "Success"])

    wb.save(path)


def test_extract_workbook_tables(tmp_path):
    workbook = tmp_path / "Parallel_MELTS_KCP-109-B_dP25.0_test.xlsx"
    _build_test_workbook(workbook)

    tables = extract_workbook_tables(workbook)
    assert "init_cond" in tables
    assert "system" in tables
    assert "liquid" in tables
    assert "quartz" in tables


def test_extract_calls_wide_and_long(tmp_path):
    workbook = tmp_path / "Parallel_MELTS_KCP-109-B_dP25.0_test.xlsx"
    _build_test_workbook(workbook)

    wide_df = extract_calls_wide(workbook)
    assert not wide_df.empty
    assert {"sample_id", "phase_name", "Index", "T (C)", "P (MPa)"}.issubset(set(wide_df.columns))
    assert {"system", "liquid"}.issubset(set(wide_df["phase_name"].unique()))

    long_df = extract_calls_long(workbook)
    assert not long_df.empty
    assert {"property_name", "property_value"}.issubset(set(long_df.columns))
    assert pd.api.types.is_numeric_dtype(long_df["property_value"])


def test_write_workbook_tables(tmp_path):
    workbook = tmp_path / "Parallel_MELTS_KCP-109-B_dP25.0_test.xlsx"
    _build_test_workbook(workbook)

    out_dir = tmp_path / "tables"
    exported = write_workbook_tables(workbook, out_dir)
    assert len(exported) >= 4
    for csv_path in exported:
        assert csv_path.exists()
        assert csv_path.suffix == ".csv"

