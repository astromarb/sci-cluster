from __future__ import annotations

import contextlib
import io
import json
import multiprocessing as mp
import os
import re
import sys
import time
import traceback
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Any

import pandas as pd
import openpyxl

sys.path.insert(0, "/Users/lopezama/PycharmProjects/sci-cluster")
import rmelts_mc_pipeline as rmp
from examples.rmelts_kcp109c_liquidus_matmul_rootcause import (
    _build_blk_cmp_from_liam_composition,
    _read_liam_init_cond,
)


LIAM_WORKBOOK = Path("/Users/lopezama/Downloads/Parallel_MELTS_KCP-109C_dP25.0_02-25_8cores.xlsx")
HELPER_DIR = Path("/Users/lopezama/Downloads")
OUT_ROOT = Path("/Users/lopezama/PycharmProjects/sci-cluster/review_outputs/rmelts_dynamic_step_audit")
PRESSURES_MPA = [400.0, 350.0, 300.0, 250.0, 225.0, 200.0, 150.0, 125.0, 100.0, 75.0, 50.0]
RUN_TIMEOUT_S = 120


@dataclass
class StepAuditResult:
    mode: str
    pressure_mpa: float
    status: str
    elapsed_s: float
    success: bool | None
    num_steps: int | None
    total_calc_time: float | None
    solidus_reached: bool | None
    liquidus_estimate_c: float | None
    first_qz_c: float | None
    first_fsp_c: float | None
    first_fsp1_c: float | None
    error_event_count: int
    timeout_event_count: int
    matmul_event_count: int
    raw_log_path: str
    steps_csv_path: str | None
    alignment_csv_path: str | None


def _now_tag() -> str:
    return time.strftime("%Y%m%d_%H%M%S")


def _note(log_path: Path, msg: str) -> None:
    line = f"[{time.strftime('%H:%M:%S')}] {msg}"
    print(line, flush=True)
    with log_path.open("a", encoding="utf-8") as f:
        f.write(line + "\n")


def _make_args(pressure_mpa: float) -> tuple[Any, ...]:
    comp, params = _read_liam_init_cond(LIAM_WORKBOOK)
    fo2_offset = float(params.get("fO2 offset", 0.0) or 0.0)
    _equil_seed, blk_cmp = _build_blk_cmp_from_liam_composition(comp, t1_c=1100.0, p_mpa=pressure_mpa, fo2_offset=fo2_offset)
    return (
        1100, 700, 1, float(pressure_mpa),
        params.get("fO2 constraint", "TRUE"),
        params.get("fO2 buffer", "NNO"),
        fo2_offset,
        0.1,
        blk_cmp,
        comp,
        "KCP109C",
        False,
    )

def _liam_system_first_temp_by_pressure(workbook_path: Path, pressure_mpa: float, tol: float = 1e-6) -> float | None:
    wb = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        if "system" not in wb.sheetnames:
            return None
        ws = wb["system"]
        header = [c for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
        try:
            p_idx = header.index("P (MPa)")
            t_idx = header.index("T (C)")
        except ValueError:
            return None
        vals = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            try:
                p = float(row[p_idx])
                t = float(row[t_idx])
            except Exception:
                continue
            if abs(p - float(pressure_mpa)) <= tol:
                vals.append(t)
        if not vals:
            return None
        return float(max(vals))
    finally:
        wb.close()


def _phase_flags(step: dict[str, Any]) -> dict[str, Any]:
    solids = step.get("solid_phases", {}) or {}
    names = sorted(list(solids.keys()))
    return {
        "solid_phase_names": ",".join(names),
        "has_quartz": "Quartz" in solids,
        "has_feldspar": "Feldspar" in solids,
        "has_feldspar_1": "Feldspar_1" in solids,
        "temperature_C": step.get("temperature"),
        "step_index": step.get("step_index"),
        "pressure_MPa": step.get("pressure"),
        "calc_time_s": step.get("calc_time"),
    }

def _workbook_phase_presence_by_tp(workbook_path: Path, phase_sheet: str, pressure_mpa: float, tol: float = 1e-6) -> set[float]:
    wb = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        if phase_sheet not in wb.sheetnames:
            return set()
        ws = wb[phase_sheet]
        header = [c for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
        try:
            p_idx = header.index("P (MPa)")
            t_idx = header.index("T (C)")
        except ValueError:
            return set()
        out: set[float] = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            try:
                p = float(row[p_idx]); t = float(row[t_idx])
            except Exception:
                continue
            if abs(p - float(pressure_mpa)) <= tol:
                out.add(float(t))
        return out
    finally:
        wb.close()


def _workbook_baseline_steps_df(workbook_path: Path, pressure_mpa: float, tol: float = 1e-6) -> pd.DataFrame:
    wb = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        ws = wb["system"]
        header = [c for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
        p_idx = header.index("P (MPa)")
        t_idx = header.index("T (C)")
        qz_t = _workbook_phase_presence_by_tp(workbook_path, "quartz", pressure_mpa, tol=tol)
        fsp_t = _workbook_phase_presence_by_tp(workbook_path, "feldspar", pressure_mpa, tol=tol)
        fsp1_t = _workbook_phase_presence_by_tp(workbook_path, "feldspar_1", pressure_mpa, tol=tol)
        rows = []
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
            try:
                p = float(row[p_idx]); t = float(row[t_idx])
            except Exception:
                continue
            if abs(p - float(pressure_mpa)) > tol:
                continue
            phases = []
            if t in qz_t:
                phases.append("Quartz")
            if t in fsp_t:
                phases.append("Feldspar")
            if t in fsp1_t:
                phases.append("Feldspar_1")
            rows.append(
                {
                    "solid_phase_names": ",".join(sorted(phases)),
                    "has_quartz": "Quartz" in phases,
                    "has_feldspar": "Feldspar" in phases,
                    "has_feldspar_1": "Feldspar_1" in phases,
                    "temperature_C": float(t),
                    "step_index": i,
                    "pressure_MPa": float(p),
                    "calc_time_s": None,
                }
            )
    finally:
        wb.close()
    df = pd.DataFrame(rows)
    if not df.empty:
        df = df.sort_values(["temperature_C"], ascending=False).reset_index(drop=True)
    return df


def _steps_dataframe(result: dict[str, Any]) -> pd.DataFrame:
    rows = []
    for step in result.get("results_data", []) or []:
        row = _phase_flags(step)
        rows.append(row)
    df = pd.DataFrame(rows)
    if not df.empty:
        df = df.sort_values(["temperature_C"], ascending=False).reset_index(drop=True)
    return df


def _first_appearance(df: pd.DataFrame, col: str) -> float | None:
    if df.empty or col not in df.columns:
        return None
    mask = df[col] == True  # noqa: E712
    if not mask.any():
        return None
    vals = pd.to_numeric(df.loc[mask, "temperature_C"], errors="coerce").dropna()
    if vals.empty:
        return None
    return float(vals.max())


def _parse_error_events(log_text: str) -> list[dict[str, Any]]:
    events: list[dict[str, Any]] = []
    pat = re.compile(r"Error at T=([0-9.]+), P=([0-9.]+): (.+)")
    for line in log_text.splitlines():
        m = pat.search(line)
        if not m:
            continue
        msg = m.group(3)
        events.append(
            {
                "temperature_C": float(m.group(1)),
                "pressure_MPa": float(m.group(2)),
                "message": msg,
                "is_timeout": "timeout" in msg.lower(),
                "is_matmul": "matmul" in msg.lower(),
            }
        )
    return events


def _worker_target(q: mp.Queue, mode: str, pressure_mpa: float, out_dir: str) -> None:
    try:
        out_path = Path(out_dir)
        run_dir = out_path / "run_local_helper" / f"{mode}_P{int(pressure_mpa)}"
        run_dir.mkdir(parents=True, exist_ok=True)

        if mode == "liam":
            helper_mod = rmp._load_helper_module(
                HELPER_DIR,
                run_dir=run_dir,
                patch_profile="profile_a_unpatched",
                patch_pressure_calc=False,
            )
        elif mode == "liam_clone":
            helper_mod = rmp._load_helper_module(
                HELPER_DIR,
                run_dir=run_dir,
                helper_implementation="liam_clone",
                patch_pressure_calc=False,
            )
        elif mode == "patched_k":
            helper_mod = rmp._load_helper_module(
                HELPER_DIR,
                run_dir=run_dir,
                helper_implementation="patched_profile_k",
                patch_pressure_calc=False,
            )
        else:
            raise ValueError(mode)

        args = _make_args(pressure_mpa)
        force_mainloop_only = os.environ.get("RMELTS_DYNAMIC_AUDIT_FORCE_LIAM_START_T", "").strip() == "1"
        forced_start_temp = None
        if force_mainloop_only:
            forced_start_temp = _liam_system_first_temp_by_pressure(LIAM_WORKBOOK, pressure_mpa)
            if forced_start_temp is None:
                raise RuntimeError(f"Could not determine Liam system start T for P={pressure_mpa} MPa")
            # Monkeypatch liquidus search to isolate the main temperature-loop behavior.
            helper_mod.find_wet_liquidus = lambda *a, **k: float(forced_start_temp)  # type: ignore[attr-defined]
        stdout_buf = io.StringIO()
        stderr_buf = io.StringIO()
        t0 = time.time()
        with contextlib.redirect_stdout(stdout_buf), contextlib.redirect_stderr(stderr_buf):
            result = helper_mod.run_single_pressure_step(args)
        elapsed = time.time() - t0
        q.put(
            {
                "status": "ok",
                "elapsed_s": elapsed,
                "result": result,
                "stdout": stdout_buf.getvalue(),
                "stderr": stderr_buf.getvalue(),
                "forced_start_temp_c": forced_start_temp,
            }
        )
    except Exception as exc:
        q.put(
            {
                "status": "error",
                "error": str(exc),
                "traceback": traceback.format_exc(),
            }
        )


def _run_guarded(mode: str, pressure_mpa: float, out_dir: Path, timeout_s: int) -> dict[str, Any]:
    q: mp.Queue = mp.Queue()
    proc = mp.Process(target=_worker_target, args=(q, mode, pressure_mpa, str(out_dir)))
    t0 = time.time()
    proc.start()
    proc.join(timeout=timeout_s)
    elapsed = time.time() - t0
    if proc.is_alive():
        proc.terminate()
        proc.join(timeout=5)
        if proc.is_alive():
            proc.kill()
            proc.join(timeout=5)
        return {"status": "timeout", "elapsed_s": elapsed, "exitcode": proc.exitcode}
    if not q.empty():
        out = q.get()
        out["outer_elapsed_s"] = elapsed
        out["exitcode"] = proc.exitcode
        return out
    return {"status": "no_result", "elapsed_s": elapsed, "exitcode": proc.exitcode}


def _write_alignment(liam_df: pd.DataFrame, patched_df: pd.DataFrame, path: Path) -> pd.DataFrame:
    all_t = sorted(
        set(pd.to_numeric(liam_df.get("temperature_C", pd.Series(dtype=float)), errors="coerce").dropna().tolist())
        | set(pd.to_numeric(patched_df.get("temperature_C", pd.Series(dtype=float)), errors="coerce").dropna().tolist()),
        reverse=True,
    )
    liam_map = {float(r["temperature_C"]): r for _, r in liam_df.iterrows()} if not liam_df.empty else {}
    patched_map = {float(r["temperature_C"]): r for _, r in patched_df.iterrows()} if not patched_df.empty else {}
    rows = []
    for t in all_t:
        lr = liam_map.get(float(t))
        pr = patched_map.get(float(t))
        rows.append(
            {
                "temperature_C": t,
                "liam_row_present": lr is not None,
                "patched_row_present": pr is not None,
                "liam_solid_phase_names": None if lr is None else lr.get("solid_phase_names"),
                "patched_solid_phase_names": None if pr is None else pr.get("solid_phase_names"),
                "liam_has_quartz": None if lr is None else bool(lr.get("has_quartz")),
                "patched_has_quartz": None if pr is None else bool(pr.get("has_quartz")),
                "liam_has_feldspar": None if lr is None else bool(lr.get("has_feldspar")),
                "patched_has_feldspar": None if pr is None else bool(pr.get("has_feldspar")),
                "liam_has_feldspar_1": None if lr is None else bool(lr.get("has_feldspar_1")),
                "patched_has_feldspar_1": None if pr is None else bool(pr.get("has_feldspar_1")),
                "same_phase_set": (lr is not None and pr is not None and str(lr.get("solid_phase_names")) == str(pr.get("solid_phase_names"))),
            }
        )
    adf = pd.DataFrame(rows)
    adf.to_csv(path, index=False)
    return adf


def main() -> None:
    tag = _now_tag()
    root = OUT_ROOT / f"kcp109c_dynamic_step_audit_{tag}"
    root.mkdir(parents=True, exist_ok=True)
    log_path = root / "workflow_notes.log"
    pressures = PRESSURES_MPA
    env_pressures = os.environ.get("RMELTS_DYNAMIC_AUDIT_PRESSURES", "").strip()
    if env_pressures:
        pressures = [float(x) for x in env_pressures.split(",") if x.strip()]
    timeout_s = int(os.environ.get("RMELTS_DYNAMIC_AUDIT_TIMEOUT_S", str(RUN_TIMEOUT_S)))

    force_mainloop_only = os.environ.get("RMELTS_DYNAMIC_AUDIT_FORCE_LIAM_START_T", "").strip() == "1"
    baseline_mode = os.environ.get("RMELTS_DYNAMIC_AUDIT_BASELINE", "liam").strip().lower()
    compare_mode = os.environ.get("RMELTS_DYNAMIC_AUDIT_COMPARE_MODE", "patched_k").strip().lower()
    if compare_mode not in {"patched_k", "liam_clone"}:
        raise ValueError(f"Unsupported RMELTS_DYNAMIC_AUDIT_COMPARE_MODE: {compare_mode}")
    _note(log_path, f"Starting dynamic step-by-step parity audit (Liam baseline vs {compare_mode}).")
    _note(log_path, f"Pressures={pressures}")
    _note(log_path, f"Per-run timeout={timeout_s}s")
    _note(log_path, f"Force main-loop-only via Liam start T={force_mainloop_only}")
    _note(log_path, f"Baseline mode={baseline_mode}")
    _note(log_path, f"Compare mode={compare_mode}")

    rows: list[StepAuditResult] = []
    geometry_rows: list[dict[str, Any]] = []

    for pressure in pressures:
        liam_run: dict[str, Any] | None = None
        liam_df = pd.DataFrame()
        liam_raw_path = str((root / "per_pressure" / f"P{int(pressure)}" / "liam_raw.log"))
        if baseline_mode == "liam":
            _note(log_path, f"Pressure {pressure} MPa: running Liam baseline direct worker")
            liam_run = _run_guarded("liam", pressure, root, timeout_s)
            _note(log_path, f"Pressure {pressure} MPa: Liam run status={liam_run.get('status')} exit={liam_run.get('exitcode')}")
        elif baseline_mode == "workbook":
            _note(log_path, f"Pressure {pressure} MPa: building Liam workbook baseline rows")
            liam_df = _workbook_baseline_steps_df(LIAM_WORKBOOK, pressure)
            _note(log_path, f"Pressure {pressure} MPa: workbook baseline rows={len(liam_df)}")
        else:
            raise ValueError(f"Unknown baseline mode: {baseline_mode}")
        _note(log_path, f"Pressure {pressure} MPa: running {compare_mode} direct worker")
        patched_run = _run_guarded(compare_mode, pressure, root, timeout_s)
        _note(log_path, f"Pressure {pressure} MPa: {compare_mode} run status={patched_run.get('status')} exit={patched_run.get('exitcode')}")

        pair_dir = root / "per_pressure" / f"P{int(pressure)}"
        pair_dir.mkdir(parents=True, exist_ok=True)

        def save_mode(mode: str, run: dict[str, Any]) -> tuple[pd.DataFrame, str]:
            raw_log_path = pair_dir / f"{mode}_raw.log"
            raw_log_text = ""
            if "stdout" in run or "stderr" in run:
                raw_log_text = (run.get("stdout") or "") + "\n" + (run.get("stderr") or "")
            elif "traceback" in run:
                raw_log_text = run.get("traceback", "")
            raw_log_path.write_text(raw_log_text, encoding="utf-8")

            if run.get("status") == "ok" and isinstance(run.get("result"), dict) and run["result"].get("success"):
                df = _steps_dataframe(run["result"])
                steps_csv = pair_dir / f"{mode}_steps.csv"
                df.to_csv(steps_csv, index=False)
                return df, str(raw_log_path)
            return pd.DataFrame(), str(raw_log_path)

        if baseline_mode == "liam":
            liam_df, liam_raw_path = save_mode("liam", liam_run or {})
        else:
            pair_dir.mkdir(parents=True, exist_ok=True)
            Path(liam_raw_path).write_text("BASELINE_FROM_WORKBOOK\n", encoding="utf-8")
            if not liam_df.empty:
                (pair_dir / "liam_steps.csv").write_text(liam_df.to_csv(index=False), encoding="utf-8")
        compare_label = compare_mode
        patched_df, patched_raw_path = save_mode(compare_label, patched_run)
        alignment_csv_path = pair_dir / f"alignment_liam_vs_{compare_label}.csv"
        alignment_df = _write_alignment(liam_df, patched_df, alignment_csv_path)

        for mode_name, run, df, raw_path in [
            ("liam", (liam_run or {"status": "ok", "forced_start_temp_c": _liam_system_first_temp_by_pressure(LIAM_WORKBOOK, pressure) if baseline_mode == "workbook" else None}), liam_df, liam_raw_path),
            (compare_label, patched_run, patched_df, patched_raw_path),
        ]:
            raw_text = Path(raw_path).read_text(encoding="utf-8", errors="ignore")
            events = _parse_error_events(raw_text)
            result = run.get("result") if run.get("status") == "ok" else {}
            liq_est = None
            if not df.empty:
                temps = pd.to_numeric(df["temperature_C"], errors="coerce").dropna()
                if not temps.empty:
                    liq_est = float(temps.max())
            rows.append(
                StepAuditResult(
                    mode=mode_name,
                    pressure_mpa=pressure,
                    status=str(run.get("status")),
                    elapsed_s=float(run.get("outer_elapsed_s", run.get("elapsed_s", 0.0)) or 0.0),
                    success=(result.get("success") if isinstance(result, dict) else None),
                    num_steps=(int(result.get("num_steps")) if isinstance(result, dict) and result.get("num_steps") is not None else None),
                    total_calc_time=(float(result.get("total_calc_time")) if isinstance(result, dict) and result.get("total_calc_time") is not None else None),
                    solidus_reached=(bool(result.get("solidus_reached")) if isinstance(result, dict) and result.get("solidus_reached") is not None else None),
                    liquidus_estimate_c=liq_est,
                    first_qz_c=_first_appearance(df, "has_quartz"),
                    first_fsp_c=_first_appearance(df, "has_feldspar"),
                    first_fsp1_c=_first_appearance(df, "has_feldspar_1"),
                    error_event_count=len(events),
                    timeout_event_count=sum(1 for e in events if e["is_timeout"]),
                    matmul_event_count=sum(1 for e in events if e["is_matmul"]),
                    raw_log_path=raw_path,
                    steps_csv_path=(str(pair_dir / f"{mode_name}_steps.csv") if not df.empty else None),
                    alignment_csv_path=(str(alignment_csv_path) if mode_name == compare_label else None),
                )
            )
            rows[-1].liquidus_estimate_c = (
                float(run.get("forced_start_temp_c"))
                if run.get("forced_start_temp_c") is not None and rows[-1].liquidus_estimate_c is None
                else rows[-1].liquidus_estimate_c
            )

        # Pairwise geometry/step divergence summary at this pressure
        geom = {
            "pressure_mpa": pressure,
            "liam_num_rows": int(len(liam_df)),
            "patched_num_rows": int(len(patched_df)),
            f"{compare_label}_num_rows": int(len(patched_df)),
            "row_count_delta": int(len(patched_df) - len(liam_df)),
            "liam_liquidus_estimate_c": (None if liam_df.empty else float(pd.to_numeric(liam_df["temperature_C"], errors="coerce").max())),
            "patched_liquidus_estimate_c": (None if patched_df.empty else float(pd.to_numeric(patched_df["temperature_C"], errors="coerce").max())),
            f"{compare_label}_liquidus_estimate_c": (None if patched_df.empty else float(pd.to_numeric(patched_df["temperature_C"], errors="coerce").max())),
            "liam_first_qz_c": _first_appearance(liam_df, "has_quartz"),
            "patched_first_qz_c": _first_appearance(patched_df, "has_quartz"),
            f"{compare_label}_first_qz_c": _first_appearance(patched_df, "has_quartz"),
            "liam_first_fsp_c": _first_appearance(liam_df, "has_feldspar"),
            "patched_first_fsp_c": _first_appearance(patched_df, "has_feldspar"),
            f"{compare_label}_first_fsp_c": _first_appearance(patched_df, "has_feldspar"),
            "liam_first_fsp1_c": _first_appearance(liam_df, "has_feldspar_1"),
            "patched_first_fsp1_c": _first_appearance(patched_df, "has_feldspar_1"),
            f"{compare_label}_first_fsp1_c": _first_appearance(patched_df, "has_feldspar_1"),
            "phase_set_mismatch_rows": int((alignment_df["same_phase_set"] == False).sum()) if not alignment_df.empty else 0,  # noqa: E712
            "liam_only_rows": int(((alignment_df["liam_row_present"] == True) & (alignment_df["patched_row_present"] == False)).sum()) if not alignment_df.empty else 0,  # noqa: E712
            "patched_only_rows": int(((alignment_df["patched_row_present"] == True) & (alignment_df["liam_row_present"] == False)).sum()) if not alignment_df.empty else 0,  # noqa: E712
        }
        for phase in ["qz", "fsp", "fsp1"]:
            lk = f"liam_first_{phase}_c"
            pk = f"{compare_label}_first_{phase}_c"
            if geom[lk] is None or geom[pk] is None:
                geom[f"delta_first_{phase}_c"] = None
            else:
                geom[f"delta_first_{phase}_c"] = float(geom[pk] - geom[lk])
        geometry_rows.append(geom)
        _note(
            log_path,
            (
                f"P={pressure:.0f} MPa rows(liam,patched)=({geom['liam_num_rows']},{geom['patched_num_rows']}) "
                f"first_fsp(liam,{compare_label})=({geom['liam_first_fsp_c']},{geom[f'{compare_label}_first_fsp_c']}) "
                f"first_qz(liam,{compare_label})=({geom['liam_first_qz_c']},{geom[f'{compare_label}_first_qz_c']})"
            ),
        )

    summary_df = pd.DataFrame([asdict(r) for r in rows]).sort_values(["pressure_mpa", "mode"])
    summary_csv = root / "dynamic_step_audit_summary.csv"
    summary_df.to_csv(summary_csv, index=False)
    geometry_df = pd.DataFrame(geometry_rows).sort_values("pressure_mpa", ascending=False)
    geometry_csv = root / "dynamic_step_geometry_compare.csv"
    geometry_df.to_csv(geometry_csv, index=False)

    report = {
        "status": "ok",
        "pressures_mpa": pressures,
        "timeout_s_per_run": timeout_s,
        "compare_mode": compare_mode,
        "summary_csv": str(summary_csv),
        "geometry_csv": str(geometry_csv),
        "results": [asdict(r) for r in rows],
        "geometry_rows": geometry_rows,
        "conclusions": {
            "all_runs_completed": bool((summary_df["status"] == "ok").all()),
            "pressures_with_fsp_shift": geometry_df.loc[geometry_df["delta_first_fsp_c"].notna() & (geometry_df["delta_first_fsp_c"] != 0), "pressure_mpa"].tolist() if not geometry_df.empty else [],
            "pressures_with_qz_shift": geometry_df.loc[geometry_df["delta_first_qz_c"].notna() & (geometry_df["delta_first_qz_c"] != 0), "pressure_mpa"].tolist() if not geometry_df.empty else [],
            "pressures_with_row_count_diff": geometry_df.loc[geometry_df["row_count_delta"] != 0, "pressure_mpa"].tolist() if not geometry_df.empty else [],
        },
    }
    report_path = root / "dynamic_step_audit_report.json"
    report_path.write_text(json.dumps(report, indent=2, default=str), encoding="utf-8")
    _note(log_path, f"Wrote dynamic audit summary: {summary_csv}")
    _note(log_path, f"Wrote geometry comparison: {geometry_csv}")
    _note(log_path, f"Wrote report: {report_path}")


if __name__ == "__main__":
    mp.set_start_method("spawn", force=True)
    try:
        main()
    except Exception:
        # Write a last-resort traceback near cwd for fast debugging if the harness crashes before report write.
        err_path = Path.cwd() / "dynamic_step_audit_unhandled_error.txt"
        err_path.write_text(traceback.format_exc(), encoding="utf-8")
        raise
