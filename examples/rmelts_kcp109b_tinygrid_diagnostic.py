from __future__ import annotations

import contextlib
import csv
import importlib.util
import io
import json
import math
import re
import sys
import time
from pathlib import Path
from typing import Any

import pandas as pd

REPO_ROOT = Path('/Users/lopezama/PycharmProjects/sci-cluster')
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

import rmelts_mc_pipeline as rmp

HELPER_DIR = Path('/Users/lopezama/Downloads')
XRF_CSV = Path('/Users/lopezama/PycharmProjects/sci-cluster/sci-data/Aplites_XRF_data/Aplites_HAL_XRF_noUncertainty.csv')
DIAG_ROOT = Path('/Users/lopezama/PycharmProjects/sci-cluster/review_outputs/rmelts_diagnostics')
SAMPLE_LABEL = 'KCP109B'

T1 = 1000
T2 = 900
DT = 1
P1 = 400
P2 = 200
DP = 100
FO2_CONSTRAINT = 'TRUE'
FO2_BUFFER = 'NNO'
FO2_OFFSET = 0.0
MAX_PRESSURE_WORKERS_REQUESTED = 11
PRESSURES = [400.0, 300.0, 200.0]

ERROR_LINE_RE = re.compile(r"Error at T=(?P<T>[-+0-9.]+), P=(?P<P>[-+0-9.]+): (?P<msg>.+)")


def _ts() -> str:
    return time.strftime('%Y%m%d_%H%M%S')


def _ensure_dir(path: Path) -> Path:
    path.mkdir(parents=True, exist_ok=True)
    return path


def _json_default(obj: Any):
    if isinstance(obj, Path):
        return str(obj)
    if hasattr(obj, 'item'):
        try:
            return obj.item()
        except Exception:
            pass
    raise TypeError(f'Not JSON serializable: {type(obj)}')


def _import_liam_helper_unpatched() -> Any:
    helper_path = HELPER_DIR / 'MeltsHelperFunctions.py'
    spec = importlib.util.spec_from_file_location('MeltsHelperFunctions', str(helper_path))
    if spec is None or spec.loader is None:
        raise RuntimeError(f'Could not import helper from {helper_path}')
    module = importlib.util.module_from_spec(spec)
    sys.modules['MeltsHelperFunctions'] = module
    spec.loader.exec_module(module)
    return module


def _build_kcp109b_prepared_csv(run_root: Path) -> dict[str, Any]:
    norm = rmp.normalize_aplite_xrf_to_rowwise_mc(
        str(XRF_CSV),
        output_dir=str(run_root / 'prep_outputs'),
        dataset_name='kcp109b_tinygrid_diag_norm',
        fixed_h2o_wt=13.0,
        target_samples=[SAMPLE_LABEL],
    )
    conv = rmp.MC_to_csv_rMELTS(
        norm.normalized_target_rowwise_csv_path or norm.normalized_all_rowwise_csv_path,
        output_dir=str(run_root / 'prep_outputs'),
        dataset_name='kcp109b_tinygrid_diag_prepared',
        sample_id_col='SampleID',
        fe_total_col='FeOt',
        fe3fet=0.0,
        h2o_col='H2O',
        validate_only=True,
    )
    return {'norm': norm, 'conv': conv}


def _load_prepared_and_wide(prepared_csv_path: str) -> tuple[pd.DataFrame, pd.DataFrame, rmp.MELTSRunParams]:
    prepared_df = pd.read_csv(prepared_csv_path)
    params = rmp.MELTSRunParams(
        T1=T1,
        T2=T2,
        dT=DT,
        P1=P1,
        P2=P2,
        dP=DP,
        fO2_constraint=FO2_CONSTRAINT,
        fO2_buffer=FO2_BUFFER,
        fO2_offset=FO2_OFFSET,
    )
    wide_df = rmp._build_melts_input_wide_dataframe(prepared_df, params)
    return prepared_df, wide_df, params


def _extract_helper_inputs_from_wide(wide_df: pd.DataFrame, sample_label: str) -> tuple[dict[str, float], dict[str, Any]]:
    cur_col = wide_df[sample_label]
    composition = pd.to_numeric(cur_col.iloc[:15], errors='coerce').fillna(0.0).to_dict()
    params = {
        'T1': float(cur_col['T1']),
        'T2': float(cur_col['T2']),
        'delta_T': float(cur_col['ΔT']),
        'P1': float(cur_col['P1']),
        'P2': float(cur_col['P2']),
        'delta_P': float(cur_col['ΔP']),
        'const_fO2': cur_col['fO2 constraint'],
        'fO2Path': cur_col['fO2 buffer'],
        'fO2_offset': float(cur_col['fO2 offset']),
    }
    return composition, params


def _build_blk_cmp_from_composition(composition: dict[str, float]):
    import numpy as np
    from thermoengine import core, model

    model_db = model.Database(liq_mod='v1.0')
    liquid = model_db.get_phase('Liq')
    elm_sys = ['H', 'O', 'Na', 'Mg', 'Al', 'Si', 'P', 'K', 'Ca', 'Ti', 'Cr', 'Mn', 'Fe', 'Co', 'Ni']
    mol_oxides = core.chem.format_mol_oxide_comp(composition, convert_grams_to_moles=True)
    moles_end, _oxide_res = liquid.calc_endmember_comp(mol_oxide_comp=mol_oxides, method='intrinsic', output_residual=True)
    if not liquid.test_endmember_comp(moles_end):
        raise RuntimeError('Calculated composition is infeasible for Liq endmember comp')
    mol_elm = liquid.convert_endmember_comp(moles_end, output='moles_elements')
    blk_cmp = np.array([mol_elm[core.chem.PERIODIC_ORDER.tolist().index(elm)] for elm in elm_sys])
    return blk_cmp


def _parse_error_lines(stdout_text: str, profile_name: str, pressure_mpa: float) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for line in stdout_text.splitlines():
        m = ERROR_LINE_RE.search(line.strip())
        if not m:
            continue
        temp = float(m.group('T'))
        msg = m.group('msg').strip()
        outcome = 'timeout' if 'timeout' in msg.lower() else 'error'
        rows.append({
            'profile': profile_name,
            'pressure_mpa': float(pressure_mpa),
            'step_index': None,
            'temperature_c': temp,
            'outcome': outcome,
            'calc_time_s': None,
            'liq_frac': None,
            'phases': '',
            'error': msg,
        })
    return rows


def _liq_frac_from_step(step_data: dict[str, Any]) -> float | None:
    try:
        system_mass = float(step_data['system_properties']['Mass'])
        liquid_mass = float(step_data['liquid_properties']['Mass'])
        water_mass = 0.0
        solid_phases = step_data.get('solid_phases') or {}
        if 'Water' in solid_phases and isinstance(solid_phases['Water'], dict):
            water_mass = float(solid_phases['Water'].get('Mass', 0.0) or 0.0)
        denom = system_mass - water_mass
        if denom <= 0:
            return None
        return liquid_mass / denom
    except Exception:
        return None


def _success_trace_rows(profile_name: str, pressure_mpa: float, result: dict[str, Any]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for step_data in result.get('results_data', []) or []:
        solids = sorted(list((step_data.get('solid_phases') or {}).keys()))
        rows.append({
            'profile': profile_name,
            'pressure_mpa': float(pressure_mpa),
            'step_index': int(step_data.get('step_index')) if step_data.get('step_index') is not None else None,
            'temperature_c': float(step_data.get('temperature')) if step_data.get('temperature') is not None else None,
            'outcome': 'success',
            'calc_time_s': float(step_data.get('calc_time')) if step_data.get('calc_time') is not None else None,
            'liq_frac': _liq_frac_from_step(step_data),
            'phases': ','.join(solids),
            'error': '',
        })
    return rows


def _print_step_rows(rows: list[dict[str, Any]], profile_name: str, pressure_mpa: float) -> None:
    # Sort by temperature descending, then success before errors at same T if possible.
    def key_fn(r):
        t = r.get('temperature_c')
        t_key = -float(t) if t is not None else float('inf')
        outcome_rank = 0 if r.get('outcome') == 'success' else 1
        step_idx = r.get('step_index') if r.get('step_index') is not None else 10**9
        return (t_key, outcome_rank, step_idx)

    for r in sorted(rows, key=key_fn):
        if r['outcome'] == 'success':
            calc_s = 'nan' if r['calc_time_s'] is None else f"{float(r['calc_time_s']):.4f}"
            liq = 'nan' if r['liq_frac'] is None else f"{float(r['liq_frac']):.4f}"
            phases = r['phases'] or '-'
            print(
                f"profile={profile_name} P_MPa={float(pressure_mpa):.1f} step={r['step_index']} "
                f"T_C={float(r['temperature_c']):.1f} outcome=success calc_s={calc_s} liq_frac={liq} phases={phases}"
            )
        else:
            print(
                f"profile={profile_name} P_MPa={float(pressure_mpa):.1f} step={r['step_index']} "
                f"T_C={float(r['temperature_c']):.1f} outcome={r['outcome']} err={r['error']}"
            )


def _write_csv(path: Path, rows: list[dict[str, Any]], fieldnames: list[str] | None = None) -> None:
    _ensure_dir(path.parent)
    if fieldnames is None:
        fieldnames = sorted({k for row in rows for k in row.keys()}) if rows else []
    with path.open('w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def _sheet_row_counts(xlsx_path: Path) -> dict[str, int]:
    if xlsx_path is None or not xlsx_path.exists() or xlsx_path.suffix.lower() not in {'.xlsx', '.xlsm', '.xltx', '.xltm'}:
        return {}
    from openpyxl import load_workbook
    try:
        wb = load_workbook(xlsx_path, read_only=True, data_only=True)
        try:
            return {name: wb[name].max_row for name in wb.sheetnames}
        finally:
            wb.close()
    except Exception:
        return {}


def _profile_module(profile_name: str, profile_dir: Path, *, patch_pressure_calc: bool = False):
    if profile_name == 'profile_a_unpatched':
        return _import_liam_helper_unpatched()
    return rmp._load_helper_module(
        HELPER_DIR,
        run_dir=profile_dir,
        patch_profile=profile_name,
        patch_pressure_calc=patch_pressure_calc,
    )


def _run_direct_worker_profile(profile_name: str, composition: dict[str, float], params: dict[str, Any], blk_cmp, profile_dir: Path) -> dict[str, Any]:
    module = _profile_module(profile_name, profile_dir, patch_pressure_calc=False)
    profile_rows: list[dict[str, Any]] = []
    pressure_summaries: list[dict[str, Any]] = []
    stdout_path = profile_dir / 'stdout_stderr_capture.log'
    full_stdout = io.StringIO()

    for pressure in PRESSURES:
        args = (
            params['T1'], params['T2'], params['delta_T'], float(pressure),
            params['const_fO2'], params['fO2Path'], params['fO2_offset'],
            0.1, blk_cmp, composition, SAMPLE_LABEL, False,
        )
        buf_out = io.StringIO()
        buf_err = io.StringIO()
        t0 = time.time()
        with contextlib.redirect_stdout(buf_out), contextlib.redirect_stderr(buf_err):
            result = module.run_single_pressure_step(args)
        wall_s = time.time() - t0
        stdout_text = buf_out.getvalue()
        stderr_text = buf_err.getvalue()
        full_stdout.write(f"=== pressure {pressure} MPa ===\n")
        full_stdout.write(stdout_text)
        if stderr_text:
            full_stdout.write("\n[stderr]\n")
            full_stdout.write(stderr_text)

        success_rows = _success_trace_rows(profile_name, pressure, result if isinstance(result, dict) else {})
        error_rows = _parse_error_lines(stdout_text + '\n' + stderr_text, profile_name, pressure)
        combined_rows = success_rows + error_rows
        profile_rows.extend(combined_rows)
        _print_step_rows(combined_rows, profile_name, pressure)

        result_steps = [r for r in success_rows if r['temperature_c'] is not None]
        temps = [float(r['temperature_c']) for r in result_steps]
        pressure_summary = {
            'profile': profile_name,
            'pressure_mpa': float(pressure),
            'num_steps': int(result.get('num_steps', 0)) if isinstance(result, dict) else 0,
            'total_calc_s': float(result.get('total_calc_time', 0.0)) if isinstance(result, dict) and result.get('total_calc_time') is not None else None,
            'wall_s': wall_s,
            'solidus_reached': bool(result.get('solidus_reached', False)) if isinstance(result, dict) else False,
            'first_T_C': max(temps) if temps else None,
            'last_T_C': min(temps) if temps else None,
            'success': bool(result.get('success', False)) if isinstance(result, dict) else False,
            'error': result.get('error', '') if isinstance(result, dict) else 'non-dict return',
            'stdout_error_lines': len(error_rows),
        }
        pressure_summaries.append(pressure_summary)
        print(
            f"profile={profile_name} P_MPa={pressure:.1f} num_steps={pressure_summary['num_steps']} "
            f"total_calc_s={(pressure_summary['total_calc_s'] if pressure_summary['total_calc_s'] is not None else float('nan')):.4f} "
            f"solidus_reached={pressure_summary['solidus_reached']} first_T={pressure_summary['first_T_C']} last_T={pressure_summary['last_T_C']}"
        )

    stdout_path.write_text(full_stdout.getvalue(), encoding='utf-8')
    _write_csv(profile_dir / f'per_step_trace_{profile_name}.csv', profile_rows,
               fieldnames=['profile','pressure_mpa','step_index','temperature_c','outcome','calc_time_s','liq_frac','phases','error'])
    _write_csv(profile_dir / f'per_pressure_summary_{profile_name}.csv', pressure_summaries)

    total_success_steps = sum(int(r.get('num_steps') or 0) for r in pressure_summaries)
    total_errors = sum(int(r.get('stdout_error_lines') or 0) for r in pressure_summaries)
    temperatures = [r['temperature_c'] for r in profile_rows if r.get('temperature_c') is not None]
    unique_t_attempts = sorted({float(t) for t in temperatures}, reverse=True)
    monotonic_ok = True
    for p in PRESSURES:
        seq = [float(r['temperature_c']) for r in profile_rows if r.get('pressure_mpa') == p and r.get('temperature_c') is not None and r.get('outcome') == 'success']
        if any(b > a for a, b in zip(seq, seq[1:])):
            monotonic_ok = False
            break

    return {
        'profile': profile_name,
        'mode': 'direct_worker',
        'total_success_steps': total_success_steps,
        'total_error_lines': total_errors,
        'num_pressure_paths': len(pressure_summaries),
        'pressure_paths_with_data': sum(1 for r in pressure_summaries if (r.get('num_steps') or 0) > 0),
        'temperature_sequence_success_monotonic_desc': monotonic_ok,
        'min_success_T_C': min([t for t in temperatures if t is not None], default=None),
        'max_success_T_C': max([t for t in temperatures if t is not None], default=None),
        'profile_dir': str(profile_dir),
        'stdout_log': str(stdout_path),
    }


def _build_comp_task(composition: dict[str, float], params: dict[str, Any], blk_cmp):
    return (SAMPLE_LABEL, composition, params, blk_cmp, MAX_PRESSURE_WORKERS_REQUESTED, False)


def _run_composition_level_profile(profile_name: str, composition: dict[str, float], params: dict[str, Any], blk_cmp, profile_dir: Path) -> dict[str, Any]:
    module = _profile_module(profile_name, profile_dir, patch_pressure_calc=True)
    comp_task = _build_comp_task(composition, params, blk_cmp)
    buf_out = io.StringIO()
    buf_err = io.StringIO()
    with rmp._pushd(profile_dir), contextlib.redirect_stdout(buf_out), contextlib.redirect_stderr(buf_err):
        t0 = time.time()
        result = module.process_single_composition_parallel(comp_task)
        wall_s = time.time() - t0
    (profile_dir / 'composition_level_stdout_stderr.log').write_text(buf_out.getvalue() + '\n[stderr]\n' + buf_err.getvalue(), encoding='utf-8')
    filename = ''
    if isinstance(result, dict):
        filename = str(result.get('filename') or '').strip()
    xlsx_path = (profile_dir / 'parallel-results' / filename) if filename else None
    sheets = _sheet_row_counts(xlsx_path)
    summary = {
        'profile': profile_name,
        'mode': 'composition_level',
        'wall_s': wall_s,
        'helper_result': result,
        'workbook_path': str(xlsx_path) if xlsx_path is not None else None,
        'sheet_row_counts': sheets,
        'num_data_points': result.get('num_data_points') if isinstance(result, dict) else None,
        'num_pressure_steps': result.get('num_pressure_steps') if isinstance(result, dict) else None,
    }
    print(
        f"profile={profile_name} manifest_status={('error' if (isinstance(result, dict) and result.get('error')) else 'success')} "
        f"num_pressure_steps={summary['num_pressure_steps']} num_data_points={summary['num_data_points']} "
        f"melts_calc_time_s={result.get('calc_time') if isinstance(result, dict) else None} "
        f"pressure_calc_time_s={result.get('pressure_calc_time') if isinstance(result, dict) else None} "
        f"total_time_s={result.get('total_time') if isinstance(result, dict) else None} sheets={list(sheets.keys())}"
    )
    return summary


def _run_pipeline_profile(profile_name: str, prepared_mc_csv_path: str, profile_dir: Path) -> dict[str, Any]:
    orig_loader = rmp._load_helper_module

    def _wrapped_loader(helper_dir, *, run_dir=None, patch_profile=None, patch_pressure_calc=True):
        return orig_loader(helper_dir, run_dir=run_dir, patch_profile=profile_name, patch_pressure_calc=patch_pressure_calc)

    rmp._load_helper_module = _wrapped_loader
    try:
        t0 = time.time()
        run = rmp.rMELTS_run(
            prepared_mc_csv_path,
            output_dir=str(profile_dir),
            dataset_name=f'pipeline_{profile_name}',
            T1=T1, T2=T2, dT=DT,
            P1=P1, P2=P2, dP=DP,
            fO2_constraint=FO2_CONSTRAINT,
            fO2_buffer=FO2_BUFFER,
            fO2_offset=FO2_OFFSET,
            max_composition_workers=1,
            max_pressure_workers=MAX_PRESSURE_WORKERS_REQUESTED,
            verbose=False,
            helper_dir=str(HELPER_DIR),
            backend='import',
        )
        wall_s = time.time() - t0
    finally:
        rmp._load_helper_module = orig_loader

    basis = rmp.rMELTS_geobarometry_basis(
        run.manifest_csv_path,
        phases=['quartz', 'feldspar', 'feldspar_1'],
        include_system=True,
        include_liquid=True,
        missing_phase_policy='skip_point',
    )
    manifest = pd.read_csv(run.manifest_csv_path)
    manifest_row = manifest.iloc[0].to_dict() if not manifest.empty else {}
    workbook_paths = [p for p in manifest.get('excel_path', pd.Series(dtype=object)).dropna().astype(str).tolist() if p]
    sheets = _sheet_row_counts(Path(workbook_paths[0])) if workbook_paths else {}
    summary = {
        'profile': profile_name,
        'mode': 'pipeline',
        'wall_s': wall_s,
        'run_dir': run.run_dir,
        'manifest_csv_path': run.manifest_csv_path,
        'manifest_row': manifest_row,
        'sheet_row_counts': sheets,
        'phase_availability_rows': int(len(basis.phase_availability_table)) if basis.phase_availability_table is not None else 0,
        'pressure_analysis_rows': int(len(basis.pressure_analysis_table)) if basis.pressure_analysis_table is not None else 0,
        'merged_long_rows': int(len(basis.merged_long_table)) if basis.merged_long_table is not None else 0,
    }
    print(
        f"profile={profile_name} manifest_status={manifest_row.get('status')} num_pressure_steps={manifest_row.get('num_pressure_steps')} "
        f"num_data_points={manifest_row.get('num_data_points')} melts_calc_time_s={manifest_row.get('melts_calc_time_s')} "
        f"pressure_calc_time_s={manifest_row.get('pressure_calc_time_s')} total_time_s={manifest_row.get('total_time_s')} sheets={list(sheets.keys())}"
    )
    return summary


def _write_json(path: Path, obj: Any) -> None:
    _ensure_dir(path.parent)
    path.write_text(json.dumps(obj, indent=2, default=_json_default), encoding='utf-8')


def main() -> int:
    run_root = _ensure_dir(DIAG_ROOT / f'kcp109b_tinygrid_patch_ab_{_ts()}')
    notes_path = run_root / 'workflow_notes.log'

    def note(msg: str) -> None:
        line = f"[{time.strftime('%H:%M:%S')}] {msg}"
        print(line)
        with notes_path.open('a', encoding='utf-8') as fh:
            fh.write(line + '\n')

    note('Starting KCP109B tiny-grid patch A/B diagnostics')
    prep_t0 = time.time()
    prep = _build_kcp109b_prepared_csv(run_root)
    prep_elapsed = time.time() - prep_t0
    note(f"Prepared normalized/prepared inputs in {prep_elapsed:.2f}s")

    prepared_df, wide_df, melts_params = _load_prepared_and_wide(prep['conv'].prepared_mc_csv_path)
    sample_label = str(prepared_df.iloc[0]['sample_label'])
    composition, helper_params = _extract_helper_inputs_from_wide(wide_df, sample_label)
    note(f"Loaded sample {sample_label} with {len(composition)} MELTS input oxides (first 15 rows)")

    blk_t0 = time.time()
    blk_cmp = _build_blk_cmp_from_composition(composition)
    blk_elapsed = time.time() - blk_t0
    note(f"Built blk_cmp in {blk_elapsed:.2f}s")

    summaries: list[dict[str, Any]] = []
    direct_profiles = [
        'profile_a_unpatched',
        'profile_b_stable_import_only',
        'profile_c_liquidus_guard_only',
        'profile_d_safe_execute_only',
        'profile_e_current_full_patch',
        'profile_f_corrected_timeout_progression',
    ]
    for profile_name in direct_profiles:
        profile_dir = _ensure_dir(run_root / profile_name / 'direct_worker')
        note(f"Running direct-worker diagnostics for {profile_name}")
        t0 = time.time()
        summary = _run_direct_worker_profile(profile_name, composition, helper_params, blk_cmp, profile_dir)
        summary['elapsed_s'] = time.time() - t0
        summaries.append(summary)
        note(f"Completed direct-worker diagnostics for {profile_name} in {summary['elapsed_s']:.2f}s")

    # Composition-level confirmation for selected profiles.
    for profile_name in ['profile_a_unpatched', 'profile_e_current_full_patch', 'profile_f_corrected_timeout_progression']:
        profile_dir = _ensure_dir(run_root / profile_name / 'composition_level')
        note(f"Running composition-level helper diagnostics for {profile_name}")
        t0 = time.time()
        summary = _run_composition_level_profile(profile_name, composition, helper_params, blk_cmp, profile_dir)
        summary['elapsed_s'] = time.time() - t0
        summaries.append(summary)
        _write_json(profile_dir / 'composition_level_summary.json', summary)
        note(f"Completed composition-level helper diagnostics for {profile_name} in {summary['elapsed_s']:.2f}s")

    # Pipeline-level confirmation only for failing and corrected profiles.
    for profile_name in ['profile_e_current_full_patch', 'profile_f_corrected_timeout_progression']:
        profile_dir = _ensure_dir(run_root / profile_name / 'pipeline_level')
        note(f"Running pipeline-level diagnostics for {profile_name}")
        t0 = time.time()
        summary = _run_pipeline_profile(profile_name, prep['conv'].prepared_mc_csv_path, profile_dir)
        summary['elapsed_s'] = time.time() - t0
        summaries.append(summary)
        _write_json(profile_dir / 'pipeline_level_summary.json', summary)
        note(f"Completed pipeline-level diagnostics for {profile_name} in {summary['elapsed_s']:.2f}s")

    _write_json(run_root / 'profile_summary.json', summaries)
    pd.DataFrame(summaries).to_csv(run_root / 'profile_summary.csv', index=False)

    # Decision-oriented diagnosis summary.
    direct_df = pd.DataFrame([s for s in summaries if s.get('mode') == 'direct_worker'])
    diagnosis = {
        'run_root': str(run_root),
        'sample_label': SAMPLE_LABEL,
        'grid': {'T1': T1, 'T2': T2, 'dT': DT, 'P1': P1, 'P2': P2, 'dP': DP},
        'profiles_tested_direct': direct_profiles,
        'normal_functioning_profiles': direct_df.loc[direct_df['pressure_paths_with_data'] > 0, 'profile'].tolist() if not direct_df.empty else [],
        'zero_data_profiles': direct_df.loc[direct_df['pressure_paths_with_data'] == 0, 'profile'].tolist() if not direct_df.empty else [],
        'direct_worker_summary': direct_df.to_dict(orient='records') if not direct_df.empty else [],
    }

    # Identify first flip from nonzero-data to zero-data in A->F order.
    first_flip = None
    seen_nonzero = False
    for p in direct_profiles:
        row = next((r for r in diagnosis['direct_worker_summary'] if r.get('profile') == p), None)
        if row is None:
            continue
        has_data = int(row.get('pressure_paths_with_data') or 0) > 0
        if has_data:
            seen_nonzero = True
        elif seen_nonzero:
            first_flip = p
            break
    diagnosis['first_flip_profile'] = first_flip

    _write_json(run_root / 'diagnosis_summary.json', diagnosis)
    note(f"Diagnostics complete. first_flip_profile={first_flip}")
    note(f"Artifacts: {run_root}")
    return 0


if __name__ == '__main__':
    raise SystemExit(main())
